import os
import glob
import datetime
from typing import Dict, Iterable, List, Optional, Tuple
from decimal import Decimal, InvalidOperation, getcontext

import openpyxl
import fnmatch


def list_source_files(folder: str, settings: Dict) -> List[str]:
    """Return a list of source files according to file_handling filters.

    Supports XLSX by default. Adds XLS and CSV if enabled.
    Respects name pattern and date filter when enabled.
    """
    patterns: List[str] = ["*.xlsx"]
    file_settings = settings.get('file_handling', {}) if settings else {}
    if file_settings.get('support_xls'):
        patterns.append("*.xls")
    # CSV support removed per request

    # Name filter
    name_filter_enabled = file_settings.get('enable_name_filter')
    name_pattern = file_settings.get('name_filter_pattern') or ""

    # Date filter
    date_filter_enabled = file_settings.get('enable_date_filter')
    date_days = int(file_settings.get('date_filter_days') or 0)
    cutoff_time: Optional[float] = None
    if date_filter_enabled and date_days > 0:
        cutoff = datetime.datetime.now() - datetime.timedelta(days=date_days)
        cutoff_time = cutoff.timestamp()

    results: List[str] = []
    for pattern in patterns:
        for path in glob.glob(os.path.join(folder, pattern)):
            base = os.path.basename(path)
            if base.startswith("~$"):
                continue
            if name_filter_enabled and name_pattern:
                if not _matches_pattern(base, name_pattern):
                    continue
            if cutoff_time is not None:
                try:
                    if os.path.getmtime(path) < cutoff_time:
                        continue
                except OSError:
                    continue
            results.append(path)

    # Duplicate handling
    # 0 = include all, 1 = keep first by name, 2 = keep latest by name
    duplicate_mode = int(file_settings.get('duplicate_action') or 0)
    if duplicate_mode in (1, 2):
        name_to_path: Dict[str, str] = {}
        name_to_time: Dict[str, float] = {}
        for p in results:
            name = os.path.splitext(os.path.basename(p))[0]
            mtime = 0.0
            try:
                mtime = os.path.getmtime(p)
            except OSError:
                pass
            if name not in name_to_path:
                name_to_path[name] = p
                name_to_time[name] = mtime
            else:
                if duplicate_mode == 1:
                    # keep first seen, do nothing
                    pass
                else:
                    # keep latest
                    if mtime >= name_to_time[name]:
                        name_to_path[name] = p
                        name_to_time[name] = mtime
        results = list(name_to_path.values())

    return sorted(results)


def _matches_pattern(filename: str, pattern: str) -> bool:
    """Case-insensitive glob-like match using fnmatch."""
    return fnmatch.fnmatch(filename.lower(), pattern.lower())


def load_cells(ws, settings: Dict) -> Iterable:
    """Yield cells from a worksheet honoring custom range and ignore_formulas."""
    data_settings = settings.get('data_processing', {}) if settings else {}
    use_range = data_settings.get('use_custom_range')
    custom_range = (data_settings.get('custom_range') or '').strip()
    ignore_formulas = bool(data_settings.get('ignore_formulas'))

    if use_range and custom_range:
        try:
            for row in ws[custom_range]:
                for cell in row:
                    if ignore_formulas and cell.data_type == 'f':
                        continue
                    yield cell
        except Exception:
            # Fall back to all cells if range invalid
            for row in ws.iter_rows():
                for cell in row:
                    if ignore_formulas and cell.data_type == 'f':
                        continue
                    yield cell
    else:
        for row in ws.iter_rows():
            for cell in row:
                if ignore_formulas and cell.data_type == 'f':
                    continue
                yield cell


def normalize_value(value, settings: Dict) -> Optional[Decimal]:
    """Convert a cell value to Decimal if possible according to data_processing settings."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Use string conversion to avoid binary float artifacts
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    data_settings = settings.get('data_processing', {}) if settings else {}
    if not data_settings.get('auto_convert_text'):
        return None

    text = str(value).strip()
    if text == "":
        return None

    # Remove currency symbols if allowed
    if data_settings.get('handle_currency'):
        for ch in "$€£₱₹":
            text = text.replace(ch, '')
        text = text.replace(',', '')

    # Handle percentages
    if data_settings.get('handle_percentages') and text.endswith('%'):
        try:
            return (Decimal(text[:-1]) / Decimal('100'))
        except (InvalidOperation, ValueError):
            return None

    # Plain number conversion
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def validate_value(val: Decimal, settings: Dict) -> bool:
    """Validate numeric value against type/range settings."""
    validation = settings.get('validation', {}) if settings else {}
    if not validation.get('validate_ranges'):
        return True
    try:
        min_raw = validation.get('min_value', float('-inf'))
        max_raw = validation.get('max_value', float('inf'))
        min_v = Decimal(str(min_raw))
        max_v = Decimal(str(max_raw))
    except (InvalidOperation, ValueError):
        # If conversion fails, fall back to permissive validation
        return True
    return (min_v <= val <= max_v)


def ensure_backup(save_folder: str, settings: Dict, consolidated_filename: str) -> Optional[str]:
    """Create backup if enabled. Returns backup path if created."""
    perf = settings.get('performance', {}) if settings else {}
    if not perf.get('create_backup'):
        return None
    backup_dir = os.path.join(save_folder, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    # timestamped backup name
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = os.path.join(backup_dir, f"{stamp}-{consolidated_filename}")
    # Source will be created later; caller should copy the file once saved.
    # Here we only maintain rotation for existing backups.
    _keep_hist = bool(perf.get('keep_backups'))
    max_backups = int(perf.get('max_backups') or 10)

    # Rotate
    existing = sorted(glob.glob(os.path.join(backup_dir, f"*-{consolidated_filename}")))
    if _keep_hist and len(existing) > max_backups:
        for old in existing[0: max(0, len(existing) - max_backups)]:
            try:
                os.remove(old)
            except OSError:
                pass

    return backup_path


