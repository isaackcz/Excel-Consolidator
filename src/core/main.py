import sys
import os
import pandas as pd
import logging
from logging.handlers import RotatingFileHandler
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QMessageBox, QFrame, QSpacerItem, QSizePolicy, QDialog, 
    QProgressBar, QGroupBox, QTextEdit, QListWidget, QSplitter, QCheckBox,
    QComboBox, QSpinBox, QLineEdit, QTabWidget, QScrollArea, QGridLayout,
    QSlider, QDoubleSpinBox, QRadioButton, QButtonGroup, QStyle
)
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor, QPixmap, QCursor
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize, QTimer, QObject, QEvent, QPoint
from PyQt5.QtGui import QHelpEvent
from typing import Optional
import webbrowser
from copy import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
import glob
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta
import json
import shutil
import hashlib
import re
import xlrd  # For .xls files
import csv
import threading
import warnings
warnings.filterwarnings('ignore')

# Import our custom modules
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))

from src.modules.google_sheets_reporter import setup_google_sheets_error_reporting, GoogleSheetsErrorReporter
from src.core.version import APP_VERSION, APP_NAME, GITHUB_OWNER, GITHUB_REPO, ERROR_REPORTING_ENABLED

# FileProcessor moved to src/utils/file_processor.py to avoid duplication
from src.utils.file_processor import FileProcessor

# ---------------- Logging Setup ----------------
def setup_processing_logger():
    """Set up comprehensive logging for Excel consolidation processing."""
    # Create logs directory if it doesn't exist
    log_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    # Create logger
    logger = logging.getLogger('excel_consolidator')
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # Create formatter
    formatter = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # File handler with rotation (max 10MB, keep 5 files)
    log_file = os.path.join(log_dir, 'consolidation_processing.log')
    file_handler = RotatingFileHandler(
        log_file, 
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    
    # Console handler (for immediate feedback)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Prevent propagation to root logger
    logger.propagate = False
    
    return logger

# Global logger instance
processing_logger = setup_processing_logger()

# ---------------- Advanced Configuration Dialog ----------------
class AdvancedConfigDialog(QDialog):
    """Advanced configuration options for consolidation"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Advanced Consolidation Settings")
        self.setModal(True)
        self.setFixedSize(600, 700)
        
        # Configure tooltip timing globally
        QApplication.instance().setAttribute(Qt.AA_DisableWindowContextHelpButton)
        self.setup_tooltip_timing()
        
        self.init_ui()
    
    def setup_tooltip_timing(self):
        """Configure tooltip display timing"""
        # Set tooltip delay to 1000ms (1 second)
        QApplication.instance().setAttribute(Qt.AA_UseHighDpiPixmaps)
        # Note: PyQt5 doesn't have direct control over tooltip fade timing
        # but we can set the show delay
        self.setToolTipDuration(5000)  # Keep tooltip visible for 5 seconds
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Create tab widget
        tabs = QTabWidget()
        
        # Data Processing Tab
        data_tab = self.create_data_processing_tab()
        tabs.addTab(data_tab, "Data Processing")
        
        # File Handling Tab
        file_tab = self.create_file_handling_tab()
        tabs.addTab(file_tab, "File Handling")
        
        # Validation Tab
        validation_tab = self.create_validation_tab()
        tabs.addTab(validation_tab, "Validation")
        
        # Performance Tab
        performance_tab = self.create_performance_tab()
        tabs.addTab(performance_tab, "Performance")
        
        layout.addWidget(tabs)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.ok_btn = QPushButton("Apply settings")
        self.ok_btn.setObjectName("SuccessButton")
        self.ok_btn.setToolTip(
            "Save all your advanced settings and use them for consolidation.\n\n"
            "After clicking this:\n"
            "• Your settings will be remembered\n"
            "• The consolidation will use these settings\n"
            "• The Advanced Settings button will show a ✓ checkmark\n\n"
            "💡 TIP: Review all tabs before applying!"
        )
        
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setObjectName("TertiaryButton")
        self.cancel_btn.setToolTip(
            "Close the Advanced Settings without saving any changes.\n\n"
            "When you click this:\n"
            "• All changes are discarded\n"
            "• Previous settings remain unchanged\n"
            "• Consolidation will use default or previous settings\n\n"
            "💡 TIP: Use this if you changed something by mistake."
        )
        
        self.reset_btn = QPushButton("Reset to defaults")
        self.reset_btn.setObjectName("DangerButton")
        self.reset_btn.setToolTip(
            "Reset ALL settings back to their original default values.\n\n"
            "This will:\n"
            "• Clear all custom settings\n"
            "• Restore factory defaults\n"
            "• Work like a fresh installation\n\n"
            "⚠️ WARNING: This cannot be undone! Make sure this is what you want."
        )
        
        btn_layout.addWidget(self.reset_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addWidget(self.ok_btn)
        
        self.ok_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)
        self.reset_btn.clicked.connect(self.reset_to_defaults)
        
        layout.addLayout(btn_layout)
    
    def create_data_processing_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Data Type Handling
        group = QGroupBox("Data Type Handling")
        group_layout = QVBoxLayout(group)
        
        self.auto_convert_text = QCheckBox("Auto-convert text numbers to numeric")
        self.auto_convert_text.setChecked(True)
        self.auto_convert_text.setToolTip(
            "When ENABLED: Automatically converts text that looks like numbers into actual numbers.\n\n"
            "Examples:\n"
            "• '123' becomes 123\n"
            "• '45.67' becomes 45.67\n"
            "• '1,234' becomes 1234\n\n"
            "When DISABLED: Only processes cells that are already numbers.\n\n"
            "💡 TIP: Keep this enabled unless your data has text that looks like numbers but shouldn't be added."
        )
        group_layout.addWidget(self.auto_convert_text)
        
        self.handle_percentages = QCheckBox("Convert percentages (e.g., '50%' → 0.5)")
        self.handle_percentages.setChecked(True)
        self.handle_percentages.setToolTip(
            "When ENABLED: Converts percentage text into decimal numbers.\n\n"
            "Examples:\n"
            "• '50%' becomes 0.5\n"
            "• '100%' becomes 1.0\n"
            "• '25%' becomes 0.25\n\n"
            "When DISABLED: Treats percentage symbols as text and ignores them.\n\n"
            "💡 TIP: Enable this if your Excel files contain percentage values as text."
        )
        group_layout.addWidget(self.handle_percentages)
        
        self.handle_currency = QCheckBox("Strip currency symbols (e.g., '$100' → 100)")
        self.handle_currency.setChecked(True)
        self.handle_currency.setToolTip(
            "When ENABLED: Removes currency symbols and converts to numbers.\n\n"
            "Examples:\n"
            "• '$100' becomes 100\n"
            "• '$1,500.50' becomes 1500.50\n"
            "• '€250' becomes 250\n\n"
            "When DISABLED: Treats currency symbols as text and ignores them.\n\n"
            "💡 TIP: Enable this if your Excel files have money amounts with $ or other currency symbols."
        )
        group_layout.addWidget(self.handle_currency)
        
        # Format Standardization (DISABLED by default - too slow!)
        self.enable_format_standardization = QCheckBox("⚠️ Pre-process source files (VERY SLOW - NOT recommended)")
        self.enable_format_standardization.setChecked(False)  # DISABLED by default
        self.enable_format_standardization.setToolTip(
            "⚠️ WARNING: This setting is VERY SLOW!\n\n"
            "When ENABLED:\n"
            "• Modifies and SAVES each source file to match template format\n"
            "• Takes ~30-60 seconds PER FILE to save\n"
            "• 10 files = 5-10 MINUTES\n"
            "• 100 files = 1-2 HOURS\n\n"
            "⚡ RECOMMENDED: Keep this DISABLED\n\n"
            "When DISABLED (FAST mode):\n"
            "• Reads files in read-only mode (no modifications)\n"
            "• Converts formats ON-THE-FLY during reading (in-memory)\n"
            "• 10 files = 10-30 SECONDS\n"
            "• 100 files = 2-5 MINUTES\n"
            "• 20x FASTER!\n\n"
            "How it works (DISABLED mode):\n"
            "• System reads values and converts them in-memory\n"
            "• No files are modified on disk\n"
            "• Template format still determines SUM vs AVERAGE\n"
            "• Same accuracy, much faster!\n\n"
            "💡 TIP: KEEP THIS DISABLED unless you specifically need to modify source files.\n\n"
            "Only enable if:\n"
            "• You want source files permanently converted to match template\n"
            "• You're willing to wait 1-2 hours for 100 files\n"
            "• Speed is not important"
        )
        group_layout.addWidget(self.enable_format_standardization)
        
        self.ignore_formulas = QCheckBox("Ignore cells with formulas (use calculated values only)")
        self.ignore_formulas.setChecked(True)
        self.ignore_formulas.setToolTip(
            "When ENABLED: Uses the calculated result of formulas, not the formula itself.\n\n"
            "Example:\n"
            "• Cell contains '=A1+B1' and shows '50'\n"
            "• System will use 50 (the result)\n\n"
            "When DISABLED: Tries to process the formula text as data.\n\n"
            "💡 TIP: Keep this enabled! You want the calculated values, not the formula text."
        )
        group_layout.addWidget(self.ignore_formulas)
        
        layout.addWidget(group)
        
        # Cell Range Selection
        range_group = QGroupBox("Cell Range Selection")
        range_layout = QVBoxLayout(range_group)
        
        self.use_custom_range = QCheckBox("Use custom cell range")
        self.use_custom_range.setToolTip(
            "When ENABLED: Only processes cells within the specified range.\n\n"
            "When DISABLED: Processes ALL cells in each Excel file.\n\n"
            "Examples of custom ranges:\n"
            "• 'A1:D10' - processes cells from A1 to D10\n"
            "• 'B2:F50' - processes cells from B2 to F50\n"
            "• 'A1:A100' - processes only column A, rows 1 to 100\n\n"
            "💡 TIP: Use this when you only want to consolidate specific areas of your spreadsheets."
        )
        range_layout.addWidget(self.use_custom_range)
        
        range_input_layout = QHBoxLayout()
        range_label = QLabel("Range:")
        range_label.setToolTip(
            "Enter the cell range you want to process.\n\n"
            "Format: StartCell:EndCell\n\n"
            "Examples:\n"
            "• A1:Z100 (columns A to Z, rows 1 to 100)\n"
            "• B5:H25 (columns B to H, rows 5 to 25)\n"
            "• A1:A1000 (only column A, rows 1 to 1000)\n\n"
            "💡 TIP: Make sure this range covers all the data you want to consolidate!"
        )
        range_input_layout.addWidget(range_label)
        
        self.range_input = QLineEdit("A1:Z100")
        self.range_input.setEnabled(False)
        self.range_input.setToolTip(
            "Type the exact cell range you want to consolidate.\n\n"
            "Must be in Excel format: StartCell:EndCell\n\n"
            "Examples:\n"
            "• A1:Z100 - processes a large area\n"
            "• B2:E20 - processes a smaller specific area\n"
            "• A1:A1000 - processes only one column\n\n"
            "⚠️ WARNING: Only cells within this range will be included in the consolidation!"
        )
        range_input_layout.addWidget(self.range_input)
        range_layout.addLayout(range_input_layout)
        
        self.use_custom_range.toggled.connect(self.range_input.setEnabled)
        
        layout.addWidget(range_group)
        
        layout.addStretch()
        return widget
    
    def create_file_handling_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # File Format Support
        format_group = QGroupBox("File Format Support")
        format_layout = QVBoxLayout(format_group)
        
        self.support_xlsx = QCheckBox("Excel 2007+ (.xlsx)")
        self.support_xlsx.setChecked(True)
        self.support_xlsx.setEnabled(False)  # Always enabled
        self.support_xlsx.setToolTip(
            "Excel 2007+ format (.xlsx files) - ALWAYS ENABLED\n\n"
            "This is the main Excel format used by:\n"
            "• Excel 2007 and newer\n"
            "• Most modern spreadsheet programs\n\n"
            "These files usually contain:\n"
            "• Advanced formatting\n"
            "• Multiple worksheets\n"
            "• Charts and images\n\n"
            "💡 INFO: This cannot be disabled because it's the primary format."
        )
        format_layout.addWidget(self.support_xlsx)
        
        self.support_xls = QCheckBox("Legacy Excel (.xls)")
        self.support_xls.setChecked(True)
        self.support_xls.setToolTip(
            "When ENABLED: Also processes older Excel files (.xls format).\n\n"
            "Legacy Excel format is used by:\n"
            "• Excel 97-2003\n"
            "• Some older business systems\n"
            "• Files created many years ago\n\n"
            "When DISABLED: Only processes modern .xlsx files.\n\n"
            "💡 TIP: Enable this if you have old Excel files mixed with new ones."
        )
        format_layout.addWidget(self.support_xls)
        
        self.support_csv = QCheckBox("CSV files (.csv)")
        self.support_csv.setChecked(True)
        self.support_csv.setToolTip(
            "When ENABLED: Also processes CSV (comma-separated values) files.\n\n"
            "CSV files are:\n"
            "• Simple text files with data separated by commas\n"
            "• Can be opened in Excel\n"
            "• Often exported from databases or other systems\n"
            "• Have no formatting, just pure data\n\n"
            "When DISABLED: Only processes Excel files.\n\n"
            "💡 TIP: Enable this if you have data exported as CSV files that need to be included."
        )
        format_layout.addWidget(self.support_csv)
        
        layout.addWidget(format_group)
        
        # Duplicate Handling
        dup_group = QGroupBox("Duplicate File Handling")
        dup_layout = QVBoxLayout(dup_group)
        
        self.duplicate_action = QComboBox()
        self.duplicate_action.addItems([
            "Skip duplicates",
            "Include all duplicates",
            "Use newest file",
            "Use largest file",
            "Prompt for each duplicate"
        ])
        self.duplicate_action.setToolTip(
            "Choose what happens when files with the same name are found:\n\n"
            "• SKIP DUPLICATES: Use only the first file found, ignore others\n"
            "• INCLUDE ALL: Process all files, even if they have the same name\n"
            "• USE NEWEST: Choose the file that was modified most recently\n"
            "• USE LARGEST: Choose the file with the biggest file size\n"
            "• PROMPT FOR EACH: Ask you to choose for every duplicate found\n\n"
            "💡 TIP: 'Use newest file' is usually the best choice for updated data."
        )
        
        dup_label = QLabel("When duplicate files are found:")
        dup_label.setToolTip(
            "Duplicate files are files that have the exact same filename.\n\n"
            "This commonly happens when:\n"
            "• Files are copied to multiple folders\n"
            "• Backup copies exist\n"
            "• Different versions of the same file exist\n\n"
            "Choose how to handle these situations."
        )
        dup_layout.addWidget(dup_label)
        dup_layout.addWidget(self.duplicate_action)
        
        layout.addWidget(dup_group)
        
        # File Filtering
        filter_group = QGroupBox("File Filtering")
        filter_layout = QVBoxLayout(filter_group)
        
        self.enable_name_filter = QCheckBox("Filter by filename pattern")
        self.enable_name_filter.setToolTip(
            "When ENABLED: Only processes files that match a specific name pattern.\n\n"
            "When DISABLED: Processes ALL files in the folder.\n\n"
            "Examples of useful patterns:\n"
            "• 'Sales_*.xlsx' - only files starting with 'Sales_'\n"
            "• '*_2024.xlsx' - only files ending with '_2024'\n"
            "• 'Report*.xlsx' - only files starting with 'Report'\n\n"
            "💡 TIP: Use this to exclude files you don't want to consolidate."
        )
        filter_layout.addWidget(self.enable_name_filter)
        
        name_filter_layout = QHBoxLayout()
        pattern_label = QLabel("Pattern:")
        pattern_label.setToolTip(
            "Enter a pattern to match filenames.\n\n"
            "Use * as a wildcard (matches anything):\n"
            "• '*.xlsx' - all Excel files\n"
            "• 'Sales_*' - files starting with 'Sales_'\n"
            "• '*_report.xlsx' - files ending with '_report.xlsx'\n"
            "• 'Q1_*_2024.xlsx' - specific pattern matching\n\n"
            "💡 TIP: Test your pattern with a few files first!"
        )
        name_filter_layout.addWidget(pattern_label)
        
        self.name_filter_pattern = QLineEdit("*.xlsx")
        self.name_filter_pattern.setEnabled(False)
        self.name_filter_pattern.setToolTip(
            "Type the filename pattern here.\n\n"
            "Wildcard rules:\n"
            "• * = matches any text\n"
            "• ? = matches any single character\n\n"
            "Examples:\n"
            "• '*.xlsx' matches: file1.xlsx, data.xlsx, report.xlsx\n"
            "• 'Sales_*' matches: Sales_Jan.xlsx, Sales_Q1.xlsx\n"
            "• '*_2024.xlsx' matches: Report_2024.xlsx, Data_2024.xlsx\n\n"
            "⚠️ WARNING: Only files matching this pattern will be processed!"
        )
        name_filter_layout.addWidget(self.name_filter_pattern)
        filter_layout.addLayout(name_filter_layout)
        
        self.enable_name_filter.toggled.connect(self.name_filter_pattern.setEnabled)
        
        # Date filtering
        self.enable_date_filter = QCheckBox("Filter by modification date")
        self.enable_date_filter.setToolTip(
            "When ENABLED: Only processes files that were modified recently.\n\n"
            "When DISABLED: Processes files regardless of when they were last changed.\n\n"
            "This is useful for:\n"
            "• Processing only today's files\n"
            "• Ignoring old, outdated files\n"
            "• Focusing on recent data only\n\n"
            "💡 TIP: Use this if you have a folder with both new and old files, but only want the recent ones."
        )
        filter_layout.addWidget(self.enable_date_filter)
        
        date_layout = QHBoxLayout()
        days_label = QLabel("Days ago:")
        days_label.setToolTip(
            "How many days back to look for files.\n\n"
            "Examples:\n"
            "• 1 day = only files from today\n"
            "• 7 days = files from the last week\n"
            "• 30 days = files from the last month\n"
            "• 365 days = files from the last year\n\n"
            "The system checks when each file was last modified."
        )
        date_layout.addWidget(days_label)
        
        self.date_filter_days = QSpinBox()
        self.date_filter_days.setRange(1, 365)
        self.date_filter_days.setValue(30)
        self.date_filter_days.setEnabled(False)
        self.date_filter_days.setToolTip(
            "Enter the number of days to look back.\n\n"
            "Examples:\n"
            "• 1 = only files modified today\n"
            "• 7 = files modified in the last week\n"
            "• 30 = files modified in the last month\n\n"
            "Range: 1 to 365 days\n\n"
            "💡 TIP: Start with 30 days and adjust based on your needs."
        )
        date_layout.addWidget(self.date_filter_days)
        filter_layout.addLayout(date_layout)
        
        self.enable_date_filter.toggled.connect(self.date_filter_days.setEnabled)
        
        layout.addWidget(filter_group)
        
        # Sheet Selection
        sheet_group = QGroupBox("Sheet Selection")
        sheet_layout = QVBoxLayout(sheet_group)
        
        self.enable_sheet_selection = QCheckBox("Select specific sheet to process")
        self.enable_sheet_selection.setToolTip(
            "When ENABLED: Allows you to choose which sheet to process from multi-sheet workbooks.\n\n"
            "When DISABLED: Uses the active sheet (default behavior).\n\n"
            "This is useful when:\n"
            "• Your template has multiple sheets but you only want to process one\n"
            "• Source files have data in different sheet names\n"
            "• You want to process a specific sheet consistently\n\n"
            "💡 TIP: Keep disabled if your files only have one sheet or you want the default behavior."
        )
        sheet_layout.addWidget(self.enable_sheet_selection)
        
        sheet_selection_layout = QHBoxLayout()
        sheet_label = QLabel("Sheet name:")
        sheet_label.setToolTip(
            "Enter the exact name of the sheet you want to process.\n\n"
            "Examples:\n"
            "• 'Sheet1' - the default first sheet\n"
            "• 'Data' - a sheet named 'Data'\n"
            "• 'Sales Report' - a sheet with spaces in the name\n\n"
            "⚠️ IMPORTANT: The sheet name must match exactly (case-sensitive).\n"
            "If the sheet doesn't exist, the active sheet will be used instead."
        )
        sheet_selection_layout.addWidget(sheet_label)
        
        self.sheet_name = QLineEdit("Sheet1")
        self.sheet_name.setEnabled(False)
        self.sheet_name.setToolTip(
            "Type the exact name of the sheet to process.\n\n"
            "Common sheet names:\n"
            "• 'Sheet1' - default first sheet\n"
            "• 'Data' - data sheet\n"
            "• 'Summary' - summary sheet\n"
            "• 'Report' - report sheet\n\n"
            "⚠️ WARNING: Sheet names are case-sensitive!\n"
            "If the sheet doesn't exist, the active sheet will be used as fallback."
        )
        sheet_selection_layout.addWidget(self.sheet_name)
        sheet_layout.addLayout(sheet_selection_layout)
        
        self.enable_sheet_selection.toggled.connect(self.sheet_name.setEnabled)
        
        layout.addWidget(sheet_group)
        
        layout.addStretch()
        return widget
    
    def create_validation_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Data Validation
        validation_group = QGroupBox("Data Validation")
        validation_layout = QVBoxLayout(validation_group)
        
        self.validate_structure = QCheckBox("Validate file structure consistency")
        self.validate_structure.setChecked(True)
        self.validate_structure.setToolTip(
            "When ENABLED: Checks that all Excel files have similar structure.\n\n"
            "This validation checks:\n"
            "• All files have the same worksheet names\n"
            "• Data appears in similar cell locations\n"
            "• Files have compatible layouts\n\n"
            "When DISABLED: Processes files without structure checking.\n\n"
            "💡 TIP: Keep this enabled to catch files that don't match your template."
        )
        validation_layout.addWidget(self.validate_structure)
        
        self.validate_data_types = QCheckBox("Validate data types")
        self.validate_data_types.setChecked(True)
        self.validate_data_types.setToolTip(
            "When ENABLED: Checks that cell values are the expected type (numbers, text, etc.).\n\n"
            "This validation catches:\n"
            "• Text in cells that should contain numbers\n"
            "• Unexpected data formats\n"
            "• Conversion issues\n\n"
            "When DISABLED: Accepts any data type without checking.\n\n"
            "💡 TIP: Keep this enabled to catch data quality issues early."
        )
        validation_layout.addWidget(self.validate_data_types)
        
        self.validate_ranges = QCheckBox("Validate value ranges")
        self.validate_ranges.setToolTip(
            "When ENABLED: Checks that numbers fall within acceptable ranges.\n\n"
            "This validation catches:\n"
            "• Numbers that are too large or too small\n"
            "• Potential data entry errors\n"
            "• Outliers that might be mistakes\n\n"
            "When DISABLED: Accepts any numeric value.\n\n"
            "💡 TIP: Enable this if you know your data should be within specific limits."
        )
        validation_layout.addWidget(self.validate_ranges)
        
        # Range validation settings
        range_layout = QHBoxLayout()
        
        min_label = QLabel("Min value:")
        min_label.setToolTip(
            "The smallest acceptable number.\n\n"
            "Any number smaller than this will be flagged as invalid.\n\n"
            "Examples:\n"
            "• 0 = no negative numbers allowed\n"
            "• -1000 = allows negative numbers down to -1000\n"
            "• 1 = only positive numbers allowed\n\n"
            "💡 TIP: Set this based on what makes sense for your data."
        )
        range_layout.addWidget(min_label)
        
        self.min_value = QDoubleSpinBox()
        self.min_value.setRange(-999999, 999999)
        self.min_value.setValue(-1000)
        self.min_value.setEnabled(False)
        self.min_value.setToolTip(
            "Enter the minimum acceptable value.\n\n"
            "Any number below this will be considered invalid.\n\n"
            "Range: -999,999 to 999,999\n\n"
            "Examples for different data types:\n"
            "• Sales amounts: 0 (no negative sales)\n"
            "• Temperature: -100 (reasonable low temperature)\n"
            "• Percentages: 0 (no negative percentages)\n"
            "• General data: -1000 (reasonable lower bound)"
        )
        range_layout.addWidget(self.min_value)
        
        max_label = QLabel("Max value:")
        max_label.setToolTip(
            "The largest acceptable number.\n\n"
            "Any number larger than this will be flagged as invalid.\n\n"
            "Examples:\n"
            "• 1000000 = allows numbers up to 1 million\n"
            "• 100 = for percentage data (0-100%)\n"
            "• 999999 = very large upper bound\n\n"
            "💡 TIP: Set this based on the maximum reasonable value for your data."
        )
        range_layout.addWidget(max_label)
        
        self.max_value = QDoubleSpinBox()
        self.max_value.setRange(-999999, 999999)
        self.max_value.setValue(1000000)
        self.max_value.setEnabled(False)
        self.max_value.setToolTip(
            "Enter the maximum acceptable value.\n\n"
            "Any number above this will be considered invalid.\n\n"
            "Range: -999,999 to 999,999\n\n"
            "Examples for different data types:\n"
            "• Sales amounts: 1000000 (1 million max)\n"
            "• Percentages: 100 (maximum 100%)\n"
            "• Counts: 9999 (reasonable upper limit)\n"
            "• General data: 1000000 (reasonable upper bound)"
        )
        range_layout.addWidget(self.max_value)
        
        validation_layout.addLayout(range_layout)
        
        self.validate_ranges.toggled.connect(self.min_value.setEnabled)
        self.validate_ranges.toggled.connect(self.max_value.setEnabled)
        
        layout.addWidget(validation_group)
        
        # Error Handling
        error_group = QGroupBox("Error Handling")
        error_layout = QVBoxLayout(error_group)
        
        self.stop_on_error = QRadioButton("Stop processing on first error")
        self.stop_on_error.setToolTip(
            "When SELECTED: Stops the entire consolidation as soon as any error is found.\n\n"
            "Use this when:\n"
            "• You want to fix errors immediately\n"
            "• Data quality is critical\n"
            "• You prefer to address issues before continuing\n\n"
            "⚠️ WARNING: If there are many errors, you'll need to fix them one by one."
        )
        
        self.continue_on_error = QRadioButton("Continue processing and report errors")
        self.continue_on_error.setChecked(True)
        self.continue_on_error.setToolTip(
            "When SELECTED: Continues processing all files even if errors are found.\n\n"
            "Benefits:\n"
            "• Processes as much data as possible\n"
            "• Shows all errors at once in the report\n"
            "• Saves time by not stopping for each error\n\n"
            "💡 TIP: This is usually the better choice - you can review all issues together."
        )
        
        error_layout.addWidget(self.stop_on_error)
        error_layout.addWidget(self.continue_on_error)
        
        layout.addWidget(error_group)
        
        layout.addStretch()
        return widget
    
    def create_performance_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Performance Settings
        perf_group = QGroupBox("Performance Optimization")
        perf_layout = QVBoxLayout(perf_group)
        
        self.enable_parallel = QCheckBox("Enable parallel processing")
        self.enable_parallel.setChecked(True)
        self.enable_parallel.setToolTip(
            "When ENABLED: Processes multiple files at the same time for faster speed.\n\n"
            "Benefits:\n"
            "• Much faster processing with many files\n"
            "• Better use of your computer's power\n"
            "• Especially helpful with 10+ files\n\n"
            "When DISABLED: Processes files one at a time.\n\n"
            "💡 TIP: Keep this enabled unless you have a very old computer."
        )
        perf_layout.addWidget(self.enable_parallel)
        
        thread_layout = QHBoxLayout()
        thread_label = QLabel("Max threads:")
        thread_label.setToolTip(
            "How many files to process at the same time.\n\n"
            "Guidelines:\n"
            "• 2-4 threads: Good for most computers\n"
            "• 4-8 threads: Good for modern computers\n"
            "• 8+ threads: Only for powerful computers\n\n"
            "💡 TIP: Higher numbers aren't always better - try 4 first."
        )
        thread_layout.addWidget(thread_label)
        
        self.max_threads = QSpinBox()
        self.max_threads.setRange(1, 16)
        self.max_threads.setValue(4)
        self.max_threads.setEnabled(True)
        self.max_threads.setToolTip(
            "Number of files to process simultaneously.\n\n"
            "Range: 1 to 16 threads\n\n"
            "Recommendations by computer type:\n"
            "• Older computer: 2 threads\n"
            "• Average computer: 4 threads\n"
            "• Fast computer: 6-8 threads\n"
            "• Very powerful computer: 8+ threads\n\n"
            "⚠️ WARNING: Too many threads can actually slow things down!"
        )
        thread_layout.addWidget(self.max_threads)
        perf_layout.addLayout(thread_layout)
        
        self.enable_parallel.toggled.connect(self.max_threads.setEnabled)
        
        self.memory_optimization = QCheckBox("Memory optimization for large files")
        self.memory_optimization.setChecked(True)
        self.memory_optimization.setToolTip(
            "When ENABLED: Uses special techniques to handle large Excel files without using too much memory.\n\n"
            "Benefits:\n"
            "• Can process very large files (100MB+)\n"
            "• Prevents 'out of memory' errors\n"
            "• Keeps your computer responsive\n\n"
            "When DISABLED: Loads entire files into memory (faster but uses more RAM).\n\n"
            "💡 TIP: Keep this enabled if you work with large Excel files."
        )
        perf_layout.addWidget(self.memory_optimization)
        
        layout.addWidget(perf_group)
        
        # Backup Settings
        backup_group = QGroupBox("Backup & Recovery")
        backup_layout = QVBoxLayout(backup_group)
        
        self.create_backup = QCheckBox("Create backup before consolidation")
        self.create_backup.setChecked(True)
        self.create_backup.setToolTip(
            "When ENABLED: Automatically saves a copy of your consolidated file before overwriting it.\n\n"
            "Benefits:\n"
            "• Protects against data loss\n"
            "• Allows you to recover previous versions\n"
            "• Peace of mind when processing important data\n\n"
            "When DISABLED: No backup is created (not recommended).\n\n"
            "💡 TIP: Always keep this enabled for safety!"
        )
        backup_layout.addWidget(self.create_backup)
        
        self.keep_backups = QCheckBox("Keep historical backups")
        self.keep_backups.setChecked(True)
        self.keep_backups.setToolTip(
            "When ENABLED: Saves multiple backup files over time.\n\n"
            "Benefits:\n"
            "• Access to previous consolidations\n"
            "• Compare results from different dates\n"
            "• Extra protection against mistakes\n\n"
            "When DISABLED: Only keeps the most recent backup.\n\n"
            "💡 TIP: Enable this if you consolidate the same files regularly."
        )
        backup_layout.addWidget(self.keep_backups)
        
        backup_count_layout = QHBoxLayout()
        backup_label = QLabel("Max backups to keep:")
        backup_label.setToolTip(
            "How many backup files to save before deleting old ones.\n\n"
            "Examples:\n"
            "• 5 backups = last 5 consolidations\n"
            "• 10 backups = last 10 consolidations\n"
            "• 20 backups = last 20 consolidations\n\n"
            "Old backups are automatically deleted to save disk space."
        )
        backup_count_layout.addWidget(backup_label)
        
        self.max_backups = QSpinBox()
        self.max_backups.setRange(1, 50)
        self.max_backups.setValue(10)
        self.max_backups.setEnabled(True)
        self.max_backups.setToolTip(
            "Number of backup files to keep.\n\n"
            "Range: 1 to 50 backups\n\n"
            "Recommended values:\n"
            "• 5 backups = good for occasional use\n"
            "• 10 backups = good for regular use (default)\n"
            "• 20+ backups = for frequent consolidations\n\n"
            "💡 TIP: 10 backups is usually enough for most people."
        )
        backup_count_layout.addWidget(self.max_backups)
        backup_layout.addLayout(backup_count_layout)
        
        self.keep_backups.toggled.connect(self.max_backups.setEnabled)
        
        layout.addWidget(backup_group)
        
        layout.addStretch()
        return widget
    
    def reset_to_defaults(self):
        """Reset all settings to defaults"""
        # Data Processing
        self.auto_convert_text.setChecked(True)
        self.handle_percentages.setChecked(True)
        self.handle_currency.setChecked(True)
        self.enable_format_standardization.setChecked(False)  # DISABLED by default - too slow!
        self.ignore_formulas.setChecked(True)
        self.use_custom_range.setChecked(False)
        self.range_input.setText("A1:Z100")
        
        # File Handling
        self.support_xls.setChecked(True)
        self.support_csv.setChecked(True)
        self.duplicate_action.setCurrentIndex(0)
        self.enable_name_filter.setChecked(False)
        self.name_filter_pattern.setText("*.xlsx")
        self.enable_date_filter.setChecked(False)
        self.date_filter_days.setValue(30)
        self.enable_sheet_selection.setChecked(False)
        self.sheet_name.setText("Sheet1")
        
        # Validation
        self.validate_structure.setChecked(True)
        self.validate_data_types.setChecked(True)
        self.validate_ranges.setChecked(False)
        self.min_value.setValue(-1000)
        self.max_value.setValue(1000000)
        self.continue_on_error.setChecked(True)
        
        # Performance
        self.enable_parallel.setChecked(True)
        self.max_threads.setValue(4)
        self.memory_optimization.setChecked(True)
        self.create_backup.setChecked(True)
        self.keep_backups.setChecked(True)
        self.max_backups.setValue(10)
    
    def get_settings(self):
        """Get all current settings as dictionary"""
        return {
            'data_processing': {
                'auto_convert_text': self.auto_convert_text.isChecked(),
                'handle_percentages': self.handle_percentages.isChecked(),
                'handle_currency': self.handle_currency.isChecked(),
                'enable_format_standardization': self.enable_format_standardization.isChecked(),
                'ignore_formulas': self.ignore_formulas.isChecked(),
                'use_custom_range': self.use_custom_range.isChecked(),
                'custom_range': self.range_input.text()
            },
            'file_handling': {
                'support_xls': self.support_xls.isChecked(),
                'support_csv': self.support_csv.isChecked(),
                'duplicate_action': self.duplicate_action.currentIndex(),
                'enable_name_filter': self.enable_name_filter.isChecked(),
                'name_filter_pattern': self.name_filter_pattern.text(),
                'enable_date_filter': self.enable_date_filter.isChecked(),
                'date_filter_days': self.date_filter_days.value(),
                'enable_sheet_selection': self.enable_sheet_selection.isChecked(),
                'sheet_name': self.sheet_name.text()
            },
            'validation': {
                'validate_structure': self.validate_structure.isChecked(),
                'validate_data_types': self.validate_data_types.isChecked(),
                'validate_ranges': self.validate_ranges.isChecked(),
                'min_value': self.min_value.value(),
                'max_value': self.max_value.value(),
                'stop_on_error': self.stop_on_error.isChecked()
            },
            'performance': {
                'enable_parallel': self.enable_parallel.isChecked(),
                'max_threads': self.max_threads.value(),
                'memory_optimization': self.memory_optimization.isChecked(),
                'create_backup': self.create_backup.isChecked(),
                'keep_backups': self.keep_backups.isChecked(),
                'max_backups': self.max_backups.value()
            }
        }
# ---------------- Modern Loading Dialog ----------------
class ModernLoadingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Processing...")
        self.setModal(True)
        self.setFixedSize(400, 200)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        layout = QVBoxLayout(self)
        
        # Loading icon
        self.loading_label = QLabel()
        self.loading_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.loading_label)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("⏳ Consolidating Excel files, please wait...")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Files processed list
        self.file_list = QListWidget()
        self.file_list.setMaximumHeight(80)
        layout.addWidget(self.file_list)
        
        self.setStyleSheet("""
            QDialog { background-color: #deeaee; }
            QLabel { color: #111827; font-size: 14px; background: transparent; }
            QProgressBar {
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                text-align: center;
                background-color: #ffffff;
                height: 18px;
            }
            QProgressBar::chunk { background-color: #10b981; border-radius: 8px; }
            QListWidget {
                background-color: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }
        """)

    def update_progress(self, value):
        self.progress_bar.setValue(value)
        
    def add_processed_file(self, filename):
        self.file_list.addItem(f"✓ {filename}")
        self.file_list.scrollToBottom()


# ---------------- Minimal Consolidation Worker (restored) ----------------
class ConsolidationWorker(QThread):
    finished = pyqtSignal(str, str)  # (status, message)
    progress = pyqtSignal(int)
    file_processed = pyqtSignal(str)

    def __init__(self, template_path, excel_folder, save_folder, settings=None, error_reporter=None, exclude_zero_percent=False):
        super().__init__()
        self.template_path = template_path
        self.excel_folder = excel_folder
        self.save_folder = save_folder
        self.settings = settings or {}
        self.error_reporter = error_reporter
        self.exclude_zero_percent = exclude_zero_percent

    def _is_percentage_format(self, format_str: str) -> bool:
        """Enhanced percentage format detection with comprehensive patterns."""
        if not format_str:
            return False
        
        format_str = str(format_str).lower()
        percentage_patterns = [
            '%', 'percent', '0.0%', '0.00%', '0%', '#,##0%', '#,##0.0%', '#,##0.00%',
            '0.0%', '0.00%', '0%', '0.0%', '0.00%', '0%', '0.0%', '0.00%',
            'general%', 'standard%', 'percentage', 'pct', 'pct%'
        ]
        
        return any(pattern in format_str for pattern in percentage_patterns)
    
    def _is_currency_format(self, format_str: str) -> bool:
        """Enhanced currency format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str)
        currency_symbols = ['$', '€', '£', '¥', '₽', '₹', '₩', '₪', '₦', '₡', '₨', '₫', '₱', '₲', '₴', '₵', '₸', '₼', '₾', '₿']
        currency_patterns = ['currency', 'money', 'dollar', 'euro', 'pound', 'yen']
        
        return (any(symbol in format_str for symbol in currency_symbols) or 
                any(pattern in format_str.lower() for pattern in currency_patterns))
    
    def _is_number_format(self, format_str: str) -> bool:
        """Enhanced number format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str)
        
        # First check if it's already identified as percentage or currency
        if self._is_percentage_format(format_str) or self._is_currency_format(format_str):
            return False
        
        number_patterns = [
            '0.00', '#,##0', '0.0', '0', '#,##0.00', '#,##0.0', '#,##0',
            'general', 'standard', 'number', 'numeric', 'decimal',
            '0.000', '0.0000', '#,##0.000', '#,##0.0000'
        ]
        
        return any(pattern in format_str for pattern in number_patterns)
    
    def _is_date_format(self, format_str: str) -> bool:
        """Enhanced date format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str).lower()
        date_patterns = [
            'mm/dd/yyyy', 'dd/mm/yyyy', 'yyyy-mm-dd', 'mm-dd-yyyy', 'dd-mm-yyyy',
            'mm/dd/yy', 'dd/mm/yy', 'yy-mm-dd', 'mm-dd-yy', 'dd-mm-yy',
            'm/d/yyyy', 'd/m/yyyy', 'm/d/yy', 'd/m/yy',
            'date', 'time', 'datetime', 'timestamp'
        ]
        
        return any(pattern in format_str for pattern in date_patterns)
    
    def _get_consolidation_method(self, format_info: dict) -> str:
        """Determine the appropriate consolidation method based on cell format."""
        if format_info.get('is_percentage', False):
            return 'average'
        elif format_info.get('is_currency', False):
            return 'sum'
        elif format_info.get('is_number', False):
            return 'sum'
        elif format_info.get('is_date', False):
            return 'sum'  # For dates, we might want to handle differently
        else:
            return 'sum'  # Default to sum for unformatted cells

    def _is_total_cell(self, cell) -> bool:
        """Detect if a cell is likely a total row/column based on common patterns."""
        if cell.value is None:
            return False
        
        # Check for common total indicators in cell value
        value_str = str(cell.value).lower().strip()
        total_indicators = ['total', 'sum', 'subtotal', 'grand total', 'totaal', 'gesamt']
        
        # Check if cell value contains total indicators
        if any(indicator in value_str for indicator in total_indicators):
            return True
        
        # Check if cell is in a row/column that might be totals based on position
        # This is a heuristic - could be enhanced based on specific needs
        return False

    def _get_user_friendly_error_message(self, error):
        """Convert technical errors into user-friendly messages with guidance."""
        error_str = str(error).lower()
        # File access errors
        if ("permission denied" in error_str or "access denied" in error_str or 
            "sharing violation" in error_str):
            return ("🔒 File Access Denied\n\n"
                   "The application cannot access one or more files. This usually happens when:\n"
                   "• Files are currently open in Excel or another program\n"
                   "• Files are in a read-only location or network drive\n"
                   "• Insufficient permissions to the folder\n"
                   "• Antivirus software is blocking file access\n\n"
                   "💡 Solution: Close all Excel files, ensure you have write permissions to the folder, "
                   "and temporarily disable real-time antivirus scanning if needed.")
        # File format errors
        elif "badzipfile" in error_str or "zipfile" in error_str:
            return ("❌ Corrupted Excel File\n\n"
                   "One or more Excel files appear to be corrupted or not valid Excel files.\n\n"
                   "💡 Solution: Try opening the problematic files in Excel first to repair them, "
                   "or exclude them from consolidation. You can also try 'File > Open and Repair' in Excel.")
        # Memory errors
        elif "memory" in error_str or "out of memory" in error_str:
            return ("💾 Insufficient Memory\n\n"
                   "The files are too large or numerous to process with available system memory.\n\n"
                   "💡 Solution: Try processing fewer files at once, close other applications to free up memory, "
                   "or restart the application to clear memory usage.")
        # Template errors
        elif "template" in error_str or "workbook" in error_str:
            return ("📋 Template File Issue\n\n"
                   "There's a problem with the template file. It may be:\n"
                   "• Corrupted or not a valid Excel file\n"
                   "• Protected with a password\n"
                   "• In an unsupported format (try saving as .xlsx)\n"
                   "• Contains complex formulas or features\n\n"
                   "💡 Solution: Use a different template file, create a new simple template, "
                   "or remove password protection.")
        # Merged cell errors (specific cases)
        elif ("mergedcell" in error_str or "read only" in error_str or 
              "read-only" in error_str or "attribute 'value' is read-only" in error_str):
            return ("🔗 Merged Cell Conflict\n\n"
                   "The template or source files contain merged cells that prevent data consolidation.\n\n"
                   "💡 What happened: Excel merged cells can only be written to at their top-left position, "
                   "but the consolidation process tried to write data to a merged cell area.\n\n"
                   "🔧 Solutions:\n"
                   "• Unmerge all cells in the template file before consolidation\n"
                   "• In Excel: Select all cells (Ctrl+A) → Home tab → Merge & Center (to unmerge)\n"
                   "• If source files have merged cells, unmerge them as well\n"
                   "• Create a new template without any merged cells")
        # Network/path errors
        elif "no such file" in error_str or "file not found" in error_str:
            return ("📄 File Not Found\n\n"
                   "One or more files specified in the consolidation could not be found.\n\n"
                   "💡 Solution: Check that all source files still exist and haven't been moved or deleted. "
                   "Refresh the file list and verify the folder path is correct.")
        # Encoding errors
        elif "encoding" in error_str or "decode" in error_str:
            return ("📝 File Encoding Issue\n\n"
                   "One or more files have text encoding problems that prevent proper reading.\n\n"
                   "💡 Solution: Try opening the files in Excel and saving them again, "
                   "or ensure they are saved with UTF-8 encoding.")
        # Disk space errors
        elif "no space" in error_str or "disk full" in error_str:
            return ("💽 Insufficient Disk Space\n\n"
                   "There is not enough disk space to create the consolidated file.\n\n"
                   "💡 Solution: Free up disk space by deleting temporary files, "
                   "or choose a different output folder with more available space.")
        # Generic fallback with the original error for debugging
        else:
            return (f"⚠️ Consolidation Error\n\n"
                   f"An unexpected error occurred during processing.\n\n"
                   f"Technical details: {str(error)}\n\n"
                   f"💡 Common solutions:\n"
                   f"• Ensure all Excel files are closed before starting\n"
                   f"• Check that files are not corrupted or password-protected\n"
                   f"• Verify you have write permissions to the output folder\n"
                   f"• Try with a smaller set of files first\n"
                   f"• Restart the application and try again")

    def _get_file_error_message(self, file_path, error):
        """Get user-friendly error message for individual file processing errors."""
        error_str = str(error).lower()
        filename = os.path.basename(file_path)
        # File access errors
        if ("permission denied" in error_str or "access denied" in error_str or 
            "sharing violation" in error_str):
            return (f"📁 File Currently Open\n\n"
                   f"The file '{filename}' is currently open in Excel or another program.\n\n"
                   f"💡 Solution: Close the file in Excel and try again. Make sure to save any changes first.")
        # File format errors
        elif "badzipfile" in error_str or "zipfile" in error_str:
            return (f"❌ Corrupted File\n\n"
                   f"The file '{filename}' is corrupted or not a valid Excel file.\n\n"
                   f"💡 Solution: Try opening the file in Excel to repair it using 'File > Open and Repair', "
                   f"or exclude this file from consolidation.")
        # Template structure errors
        elif "template" in error_str or "structure" in error_str:
            return (f"📋 Structure Mismatch\n\n"
                   f"The file '{filename}' has a different structure than the template.\n\n"
                   f"💡 Solution: Ensure all files have the same column headers and data layout as the template.")
        # Memory errors
        elif "memory" in error_str:
            return (f"💾 File Too Large\n\n"
                   f"The file '{filename}' is too large to process with available memory.\n\n"
                   f"💡 Solution: Close other applications to free up memory, or exclude this file and process it separately.")
        # Password protected files
        elif "password" in error_str or "encrypted" in error_str:
            return (f"🔒 Password Protected\n\n"
                   f"The file '{filename}' is password protected and cannot be opened.\n\n"
                   f"💡 Solution: Remove the password protection from the file before consolidation.")
        # Network/path errors
        elif "no such file" in error_str or "file not found" in error_str:
            return (f"📄 File Not Found\n\n"
                   f"The file '{filename}' could not be found at the specified location.\n\n"
                   f"💡 Solution: Check that the file still exists and hasn't been moved or deleted.")
        # Merged cell specific errors
        elif ("mergedcell" in error_str or "attribute 'value' is read-only" in error_str or 
              "read-only" in error_str):
            return (f"🔗 Merged Cells in File\n\n"
                   f"The file '{filename}' contains merged cells that prevent data consolidation.\n\n"
                   f"💡 Solution: Open '{filename}' in Excel, select all cells (Ctrl+A), "
                   f"then click 'Merge & Center' in the Home tab to unmerge all cells. "
                   f"Save the file and try consolidation again.")
        # Generic file error
        else:
            return (f"⚠️ Processing Error\n\n"
                   f"The file '{filename}' could not be processed.\n\n"
                   f"Technical details: {str(error)}\n\n"
                   f"💡 Common solutions:\n"
                   f"• Ensure the file is not open in Excel\n"
                   f"• Check that the file is not corrupted\n"
                   f"• Verify the file format matches other files\n"
                   f"• Try excluding this file if the issue persists")

    def _get_worksheet(self, workbook, file_type="source"):
        """Get the appropriate worksheet based on settings."""
        file_handling = self.settings.get('file_handling', {})
        if file_handling.get('enable_sheet_selection', False):
            sheet_name = file_handling.get('sheet_name', 'Sheet1')
            if sheet_name in workbook.sheetnames:
                return workbook[sheet_name]
            return workbook.active
        return workbook.active
    
    def _process_cell_value_with_format_verification(self, value, format_info, coord, file_label, wb, stop_on_error):
        """
        Process cell value with comprehensive format verification.
        Ensures data is converted according to template format requirements.
        
        Args:
            value: The cell value to process
            format_info: Dictionary containing format information from template
            coord: Cell coordinate (e.g., 'A1')
            file_label: Name of the source file
            wb: Workbook object
            stop_on_error: Whether to stop on format mismatches
            
        Returns:
            Decimal value or None if processing failed
        """
        if value is None or value == "":
            return None
            
        # Handle different format types based on template requirements
        if format_info.get('is_percentage', False):
            return self._process_percentage_value(value, coord, file_label, wb, stop_on_error)
        elif format_info.get('is_currency', False):
            return self._process_currency_value(value, coord, file_label, wb, stop_on_error)
        elif format_info.get('is_number', False):
            return self._process_number_value(value, coord, file_label, wb, stop_on_error)
        else:
            # Default processing for unformatted cells
            return self._process_default_value(value, coord, file_label, wb, stop_on_error)
    
    def _process_percentage_value(self, value, coord, file_label, wb, stop_on_error):
        """Process percentage values with strict format verification."""
        try:
            # Handle different percentage input formats
            if isinstance(value, (int, float)):
                # Normalize numeric inputs to PERCENT POINTS for averaging
                # Rules:
                #  - Values > 1 are treated as percent points (e.g., 82.5 means 82.5%)
                #  - Values between 0 and 1 are decimals; convert to percent points (0.825 → 82.5)
                numeric_val = float(value)
                if 0 <= numeric_val <= 1:
                    normalized = numeric_val * 100.0
                else:
                    normalized = numeric_val
                return Decimal(str(normalized))
            elif isinstance(value, str):
                text = str(value).strip().replace(",", "")
                if text.endswith('%'):
                    # Remove % and interpret as percent points directly
                    val = Decimal(text[:-1])
                    return val
                else:
                    # Parse as number; apply same normalization as numeric path
                    numeric_val = float(text)
                    if 0 <= numeric_val <= 1:
                        normalized = numeric_val * 100.0
                    else:
                        normalized = numeric_val
                    return Decimal(str(normalized))
            else:
                return None
        except Exception as e:
            if stop_on_error:
                filename = os.path.basename(file_label) if hasattr(file_label, '__iter__') else str(file_label)
                error_msg = (f"Percentage Format Error\n\n"
                           f"Cell {coord} in file '{filename}' contains invalid percentage data:\n"
                           f"'{value}'\n\n"
                           f"Expected: Numeric values or percentages (e.g., 100, 0.5, 0.75)\n\n"
                           f"💡 Solution: Ensure the cell contains valid percentage data or "
                           f"convert the template cell to a different format.")
                self.finished.emit("error", error_msg)
            return None
    
    def _process_currency_value(self, value, coord, file_label, wb, stop_on_error):
        """Process currency values with format verification."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Remove currency symbols and parse
                text = str(value).strip().replace("$", "").replace("€", "").replace("£", "").replace("¥", "").replace(",", "")
                return Decimal(text)
            else:
                return None
        except Exception as e:
            if stop_on_error:
                filename = os.path.basename(file_label) if hasattr(file_label, '__iter__') else str(file_label)
                error_msg = (f"Currency Format Error\n\n"
                           f"Cell {coord} in file '{filename}' contains invalid currency data:\n"
                           f"'{value}'\n\n"
                           f"Expected: Numeric values (e.g., 100, 100.50)\n\n"
                           f"💡 Solution: Ensure the cell contains valid numeric data.")
                self.finished.emit("error", error_msg)
            return None
    
    def _process_number_value(self, value, coord, file_label, wb, stop_on_error):
        """Process number values with format verification."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Remove common formatting characters
                text = str(value).strip().replace(",", "").replace(" ", "")
                return Decimal(text)
            else:
                return None
        except Exception as e:
            if stop_on_error:
                filename = os.path.basename(file_label) if hasattr(file_label, '__iter__') else str(file_label)
                error_msg = (f"Number Format Error\n\n"
                           f"Cell {coord} in file '{filename}' contains invalid numeric data:\n"
                           f"'{value}'\n\n"
                           f"Expected: Numeric values (e.g., 100, 100.50)\n\n"
                           f"💡 Solution: Ensure the cell contains valid numeric data.")
                self.finished.emit("error", error_msg)
            return None
    
    def _process_default_value(self, value, coord, file_label, wb, stop_on_error):
        """Process values with default (unformatted) handling."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Try to parse as number
                text = str(value).strip().replace(",", "")
                return Decimal(text)
            else:
                return None
        except Exception:
            return None
    
    def _update_submitted_files_format(self, files, coord_format_info):
        """
        Update all submitted files to match template cell formats before consolidation.
        This ensures uniform formatting and correct processing (sum vs average).
        
        ENHANCED CONFLICT RESOLUTION:
        - Converts text to numbers when template expects numbers (SUM)
        - Converts text/numbers to percentages when template expects percentages (AVG)
        - Applies template format as the source of truth
        - Handles all edge cases: "100", "50%", 0.5, etc.
        
        OPTIMIZED FOR PERFORMANCE:
        - Uses data_only=True for faster loading
        - Only processes cells that need conversion
        - Early exit for unchanged cells
        - Progress reporting every file
        """
        processing_logger.info(f"🔧 Starting format standardization for {len(files)} files...")
        processing_logger.info(f"📊 Template format info: {len(coord_format_info)} coordinates")
        
        files_updated = 0
        cells_converted = 0
        total_files = len(files)
        
        try:
            for file_idx, file in enumerate(files, 1):
                try:
                    # Update progress for each file
                    file_progress = int((file_idx / total_files) * 100)
                    processing_logger.info(f"⚡ Processing file {file_idx}/{total_files} ({file_progress}%): {os.path.basename(file)}")
                    
                    ext = os.path.splitext(file)[1].lower()
                    if ext not in ('.xlsx', '.xls'):
                        processing_logger.info(f"  ⏩ Skipped (not Excel file)")
                        continue
                    
                    # OPTIMIZED: Load with data_only=True for speed, then reload for writing only if needed
                    # First pass: Check if any cells need conversion (read-only, fast)
                    wb_check = openpyxl.load_workbook(file, data_only=True, read_only=True)
                    ws_check = self._get_worksheet(wb_check, "source")
                    
                    cells_needing_conversion = []
                    
                    # Quick scan: identify cells that need conversion
                    for coord, format_info in coord_format_info.items():
                        if coord not in ws_check:
                            continue
                        
                        cell = ws_check[coord]
                        
                        # Skip empty cells
                        if cell.value is None or cell.value == '':
                            continue
                        
                        # Skip if already correct type
                        if self._cell_already_correct_format(cell.value, format_info):
                            continue
                        
                        cells_needing_conversion.append(coord)
                    
                    wb_check.close()
                    
                    # If no cells need conversion, skip this file entirely
                    if not cells_needing_conversion:
                        processing_logger.info(f"  ✅ No conversion needed (already matches template)")
                        continue
                    
                    processing_logger.info(f"  🔧 {len(cells_needing_conversion)} cells need conversion")
                    
                    # Second pass: Only reload for writing if conversions are needed
                    wb = openpyxl.load_workbook(file, data_only=False)
                    ws = self._get_worksheet(wb, "source")
                    
                    file_cells_updated = 0
                    
                    # OPTIMIZED: Only process cells that need conversion
                    for coord in cells_needing_conversion:
                        if coord not in ws:
                            continue
                            
                        cell = ws[coord]
                        format_info = coord_format_info[coord]
                        
                        # CRITICAL: Preserve formulas - never modify cells with formulas
                        if self._preserve_formulas_during_format_update(cell, format_info, coord):
                            continue
                        
                        original_value = cell.value
                        converted = False
                        
                        # === PERCENTAGE FORMAT (Template expects %, will use AVG) ===
                        if format_info.get('is_percentage', False):
                            converted_value = self._convert_to_percentage_format(original_value, coord)
                            if converted_value is not None:
                                cell.value = converted_value
                                cell.number_format = format_info.get('number_format', '0.00%')
                                converted = True
                        
                        # === CURRENCY FORMAT (Template expects currency, will use SUM) ===
                        elif format_info.get('is_currency', False):
                            converted_value = self._convert_to_number_format(original_value, coord, is_currency=True)
                            if converted_value is not None:
                                cell.value = converted_value
                                cell.number_format = format_info.get('number_format', '$#,##0.00')
                                converted = True
                        
                        # === NUMBER FORMAT (Template expects number, will use SUM) ===
                        elif format_info.get('is_number', False):
                            converted_value = self._convert_to_number_format(original_value, coord, is_currency=False)
                            if converted_value is not None:
                                cell.value = converted_value
                                cell.number_format = format_info.get('number_format', '#,##0.00')
                                converted = True
                        
                        # === DEFAULT (Apply template format, will use SUM) ===
                        else:
                            template_format = format_info.get('number_format')
                            if template_format:
                                cell.number_format = template_format
                                # Try to convert text to number if it looks numeric
                                if isinstance(original_value, str):
                                    converted_value = self._convert_to_number_format(original_value, coord, is_currency=False)
                                    if converted_value is not None:
                                        cell.value = converted_value
                                        converted = True
                        
                        if converted:
                            file_cells_updated += 1
                            cells_converted += 1
                    
                    # Save the updated file if any cells were converted
                    if file_cells_updated > 0:
                        processing_logger.info(f"  💾 Saving {file_cells_updated} changes...")
                        wb.save(file)
                        files_updated += 1
                        processing_logger.info(f"  ✅ Saved successfully")
                    
                    wb.close()
                    
                except Exception as e:
                    # Log error but continue with other files
                    processing_logger.warning(f"⚠️ Could not update format for {os.path.basename(file)}: {e}")
                    continue
                    
        except Exception as e:
            processing_logger.error(f"❌ Error during format standardization: {e}")
            # Continue with consolidation even if format update fails
        
        processing_logger.info(f"✅ Format standardization completed: {files_updated} files, {cells_converted} cells converted")
    
    def _cell_already_correct_format(self, value, format_info):
        """
        Quick check if cell value already matches expected format.
        Used to skip unnecessary conversions for performance.
        """
        if value is None or value == '':
            return True  # Empty cells don't need conversion
        
        # Check if numeric value already matches percentage format
        if format_info.get('is_percentage', False):
            if isinstance(value, (int, float)) and 0 <= value <= 1:
                return True  # Already in decimal format (0.825 for 82.5%)
        
        # Check if already numeric for number/currency formats
        elif format_info.get('is_number', False) or format_info.get('is_currency', False):
            if isinstance(value, (int, float)):
                return True  # Already a number
        
        return False  # Needs conversion
    
    def _convert_to_percentage_format(self, value, coord):
        """
        Convert any value to percentage format (decimal for Excel).
        Handles: numbers (82.5), decimals (0.825), text ("82.5%", "50"), etc.
        
        Returns: Decimal value for Excel (e.g., 0.825 for 82.5%)
        """
        try:
            # If already a number
            if isinstance(value, (int, float)):
                # Values > 1 are percentage points (82.5 means 82.5%)
                if value > 1:
                    return value / 100  # 82.5 → 0.825
                # Values 0-1 are already decimals (0.825 means 82.5%)
                elif 0 <= value <= 1:
                    return value  # 0.825 → 0.825
                else:
                    # Negative or unusual values, treat as percentage points
                    return value / 100
            
            # If text, parse it
            elif isinstance(value, str):
                text = value.strip()
                
                # Remove % symbol if present
                if text.endswith('%'):
                    # "82.5%" → 82.5 → 0.825
                    numeric = float(text[:-1].replace(',', ''))
                    return numeric / 100
                else:
                    # "82.5" or "0.825" - determine which
                    numeric = float(text.replace(',', ''))
                    if numeric > 1:
                        return numeric / 100  # 82.5 → 0.825
                    else:
                        return numeric  # 0.825 → 0.825
            
            return None
            
        except Exception as e:
            processing_logger.warning(f"⚠️ Could not convert {coord} value '{value}' to percentage: {e}")
            return None
    
    def _convert_to_number_format(self, value, coord, is_currency=False):
        """
        Convert any value to number format.
        Handles: numbers (100), text ("100", "$100", "1,234"), etc.
        
        Returns: Numeric value
        """
        try:
            # If already a number, return as-is
            if isinstance(value, (int, float)):
                return value
            
            # If text, parse it
            elif isinstance(value, str):
                text = value.strip()
                
                # Remove currency symbols
                for symbol in ['$', '€', '£', '¥', '₽', '₹', '₩', '₪', '₦', '₡', '₨', '₫', '₱']:
                    text = text.replace(symbol, '')
                
                # Remove commas and spaces
                text = text.replace(',', '').replace(' ', '')
                
                # Remove % symbol if present (shouldn't be here, but handle it)
                if text.endswith('%'):
                    text = text[:-1]
                
                # Parse to number
                return float(text)
            
            return None
            
        except Exception as e:
            processing_logger.warning(f"⚠️ Could not convert {coord} value '{value}' to number: {e}")
            return None

    def _validate_cell_format_consistency(self, cell, format_info, coord):
        """
        Validate that cell value is consistent with detected format.
        This helps ensure proper format detection and processing.
        """
        try:
            value = cell.value
            if value is None:
                return
            
            # Validate percentage format consistency
            if format_info.get('is_percentage', False):
                if isinstance(value, (int, float)):
                    # Check if value is in reasonable percentage range
                    if value > 1 and value <= 100:
                        # Likely percentage value (50, 75, etc.)
                        print(f"  📊 {coord}: Percentage value {value}% detected")
                    elif 0 <= value <= 1:
                        # Likely decimal percentage (0.5, 0.75, etc.)
                        print(f"  📊 {coord}: Decimal percentage {value} detected")
                    else:
                        print(f"  ⚠️ {coord}: Unusual percentage value {value}")
            
            # Validate currency format consistency
            elif format_info.get('is_currency', False):
                if isinstance(value, (int, float)):
                    print(f"  💰 {coord}: Currency value {value} detected")
            
            # Validate number format consistency
            elif format_info.get('is_number', False):
                if isinstance(value, (int, float)):
                    print(f"  🔢 {coord}: Number value {value} detected")
                    
        except Exception as e:
            print(f"  ⚠️ Format validation error for {coord}: {e}")

    def _preserve_formulas_during_format_update(self, cell, format_info, coord):
        """
        Ensure formulas are completely preserved during format updates.
        This is critical to prevent damage to existing formulas.
        """
        try:
            # Check if cell contains a formula
            has_formula = False
            formula_text = None
            
            # Multiple ways to detect formulas
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                has_formula = True
                formula_text = getattr(cell, 'formula', None)
            elif isinstance(cell.value, str) and str(cell.value).startswith('='):
                has_formula = True
                formula_text = str(cell.value)
            
            if has_formula:
                print(f"  🔒 Preserving formula in {coord}: {formula_text}")
                # Mark this cell as having a formula in format info
                format_info['has_formula'] = True
                format_info['formula_text'] = formula_text
                return True  # Indicates formula was found and preserved
            
            return False  # No formula found
            
        except Exception as e:
            print(f"  ⚠️ Formula preservation error for {coord}: {e}")
            return False

    def _get_template_load_error_message(self, error):
        """Get user-friendly error message for template loading errors."""
        error_str = str(error).lower()
        filename = os.path.basename(self.template_path)
        # Password protected files
        if "password" in error_str or "encrypted" in error_str:
            return (f"🔒 Password Protected Template\n\n"
                   f"The template file '{filename}' is password protected and cannot be opened.\n\n"
                   f"💡 Solution: Remove the password protection from the template file before using it. "
                   f"In Excel, go to File > Info > Protect Workbook > Encrypt with Password and remove the password.")
        # Corrupted template
        elif "badzipfile" in error_str or "zipfile" in error_str:
            return (f"❌ Corrupted Template File\n\n"
                   f"The template file '{filename}' appears to be corrupted or damaged.\n\n"
                   f"💡 Solution: Try opening the file in Excel to repair it using 'File > Open and Repair', "
                   f"or use a different template file. You may also try creating a new template.")
        # File format issues
        elif "invalid" in error_str or "format" in error_str:
            return (f"📄 Invalid Template Format\n\n"
                   f"The template file '{filename}' is not a valid Excel file or is in an unsupported format.\n\n"
                   f"💡 Solution: Ensure the file is a .xlsx or .xlsm file created by Excel. "
                   f"Avoid using .xls (older format) or files from other spreadsheet applications.")
        # Permission issues
        elif ("permission" in error_str or "access" in error_str or 
              "sharing violation" in error_str):
            return (f"🔐 Template Access Denied\n\n"
                   f"Cannot access the template file '{filename}'. It may be open in another program.\n\n"
                   f"💡 Solution: Ensure the template file is not open in Excel or another program, "
                   f"and verify you have read permissions to the file location.")
        # File not found
        elif "no such file" in error_str or "file not found" in error_str:
            return (f"📄 Template File Not Found\n\n"
                   f"The template file '{filename}' could not be found at the specified location.\n\n"
                   f"💡 Solution: Check that the template file still exists and hasn't been moved or deleted. "
                   f"You may need to select a new template file.")
        # Generic template error
        else:
            return (f"📋 Template Loading Error\n\n"
                   f"Could not load the template file '{filename}'.\n\n"
                   f"Technical details: {str(error)}\n\n"
                   f"💡 Solutions to try:\n"
                   f"• Open the file in Excel first to ensure it's valid\n"
                   f"• Create a new simple template with just headers\n"
                   f"• Check that the file is not corrupted or password-protected\n"
                   f"• Try using a different template file")

    def _get_save_error_message(self, error, output_path):
        """Get user-friendly error message for file saving errors."""
        error_str = str(error).lower()
        filename = os.path.basename(output_path)
        folder = os.path.dirname(output_path)
        # Permission denied
        if ("permission denied" in error_str or "access denied" in error_str or 
            "sharing violation" in error_str):
            return (f"🔐 Cannot Save File\n\n"
                   f"The consolidated file '{filename}' is currently open in Excel or another program.\n\n"
                   f"💡 Solution: Close the file in Excel and try again. Make sure to save any changes first. "
                   f"If the file appears closed, wait a moment and try again as Excel may still be releasing the file.")
        # File in use
        elif ("being used" in error_str or "in use" in error_str or 
              "process cannot access" in error_str):
            return (f"📁 File Currently Open\n\n"
                   f"The consolidated file '{filename}' is currently open in Excel or another program.\n\n"
                   f"💡 Solution: Close the file in Excel and try again. Make sure to save any changes first. "
                   f"If the file appears closed, wait a moment and try again as Excel may still be releasing the file.")
        # Disk space
        elif "no space" in error_str or "disk full" in error_str:
            return (f"💽 Insufficient Disk Space\n\n"
                   f"Not enough disk space to save the consolidated file '{filename}'.\n\n"
                   f"💡 Solution: Free up disk space by deleting temporary files and downloads, "
                   f"or choose a different output folder on a drive with more available space.")
        # Path too long
        elif "path too long" in error_str or "filename too long" in error_str:
            return (f"📏 Path Too Long\n\n"
                   f"The file path is too long for the system to handle.\n\n"
                   f"💡 Solution: Choose a shorter folder path or filename. "
                   f"Try saving to a folder closer to the root of your drive (e.g., C:\\Consolidated\\).")
        # Read-only file system
        elif "read-only" in error_str or "readonly" in error_str:
            return (f"🔒 Read-Only Location\n\n"
                   f"Cannot save to '{folder}' because it is read-only.\n\n"
                   f"💡 Solution: Choose a different output folder where you have write permissions, "
                   f"such as your Documents folder or Desktop.")
        # Network errors
        elif "network" in error_str or "unc" in error_str:
            return (f"🌐 Network Save Error\n\n"
                   f"Cannot save '{filename}' to the network location.\n\n"
                   f"💡 Solution: Check your network connection, ensure the network path is accessible, "
                   f"or try saving to a local folder first.")
        # Generic save error
        else:
            return (f"💾 Save Error\n\n"
                   f"Could not save the consolidated file '{filename}'.\n\n"
                   f"Technical details: {str(error)}\n\n"
                   f"💡 Solutions to try:\n"
                   f"• Check that the output folder exists and you have write permissions\n"
                   f"• Ensure the file is not already open in Excel\n"
                   f"• Try saving to a different location\n"
                   f"• Close other applications to free up resources\n"
                   f"• Restart the application and try again")

    def run(self):
        try:
            processing_logger.info("🚀 Starting Excel consolidation process...")
            processing_logger.info(f"📋 Template: {self.template_path}")
            processing_logger.info(f"📁 Source folder: {self.excel_folder}")
            processing_logger.info(f"💾 Output folder: {self.save_folder}")
            self.progress.emit(5)
            if not os.path.exists(self.template_path):
                error_msg = (f"📋 Template File Not Found\n\n"
                           f"The template file could not be found at:\n"
                           f"'{self.template_path}'\n\n"
                           f"💡 Please check:\n"
                           f"• The file path is correct and hasn't changed\n"
                           f"• The file still exists and hasn't been moved or deleted\n"
                           f"• You have read permissions to access the file\n"
                           f"• The file is not currently open in Excel\n"
                           f"• The network connection is stable (if on a network drive)\n\n"
                           f"🔧 Solutions:\n"
                           f"• Browse and select the template file again\n"
                           f"• Create a new template file with the required structure\n"
                           f"• Copy the template to a local folder for better access")
                self.finished.emit("error", error_msg)
                return

            template_ext = os.path.splitext(self.template_path)[1].lower()
            keep_vba = template_ext == '.xlsm'
            try:
                output_wb = openpyxl.load_workbook(self.template_path, keep_vba=keep_vba)
                output_ws = self._get_worksheet(output_wb, "template")
            except Exception as e:
                if self.error_reporter:
                    try:
                        self.error_reporter.report_error(
                            type(e), e, e.__traceback__,
                            triggered_by="Template Loading in ConsolidationWorker",
                            user_file=self.template_path
                        )
                    except Exception:
                        pass
                error_msg = self._get_template_load_error_message(e)
                self.finished.emit("error", error_msg)
                return

            try:
                from src.modules.advanced_settings import list_source_files, load_cells, normalize_value, validate_value, ensure_backup
            except Exception as e:
                list_source_files = None
                load_cells = None
                normalize_value = None
                validate_value = None
                ensure_backup = None
                # Advanced settings module not available - use basic functionality
                if self.error_reporter:
                    try:
                        self.error_reporter.report_error(
                            type(e), e, e.__traceback__,
                            triggered_by="Advanced Settings Import",
                            user_file=None
                        )
                    except Exception:
                        pass

            if list_source_files is not None:
                files = list_source_files(self.excel_folder, self.settings)
            else:
                pattern = os.path.join(self.excel_folder, "*.xlsx")
                files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]
            if not files:
                error_msg = ("📁 No Excel Files Found\n\n"
                           f"No valid Excel files were found in the folder:\n"
                           f"'{self.excel_folder}'\n\n"
                           f"💡 Please check:\n"
                           f"• The folder path is correct and accessible\n"
                           f"• The folder contains .xlsx or .xlsm files\n"
                           f"• Files are not hidden or located in subfolders\n"
                           f"• Files are not currently open in Excel (temporary files start with ~$)\n"
                           f"• You have read permissions to the folder\n\n"
                           f"📝 Supported file types: .xlsx, .xlsm\n"
                           f"❌ Not supported: .xls (older Excel format), .csv")
                self.finished.emit("error", error_msg)
                return

            totals = {}
            contributions = {}
            percent_counts = {}
            coord_is_percent = {}
            # Maintain a stable, complete list of all file labels for reporting
            all_file_labels = [os.path.splitext(os.path.basename(p))[0] for p in files]
            all_file_labels.sort(key=lambda n: n.lower())
            total_files_count = len(files)  # Total number of files for accurate counting
            validation_settings = self.settings.get('validation', {})
            validate_structure = bool(validation_settings.get('validate_structure'))
            validate_data_types = bool(validation_settings.get('validate_data_types'))
            stop_on_error = bool(validation_settings.get('stop_on_error'))

            template_ws = None
            coord_format_info = {}  # Enhanced format information storage
            template_coords = None
            
            # Enhanced template format analysis with comprehensive cell format verification
            try:
                processing_logger.info("🔍 Starting template format analysis...")
                template_wb = openpyxl.load_workbook(self.template_path, data_only=False, read_only=False)
                template_ws = self._get_worksheet(template_wb, "template")
                processing_logger.info(f"📋 Template worksheet loaded: {template_ws.title}")
                
                # Build comprehensive format cache with detailed cell format information
                coord_format_info = {}
                template_coords = set()
                
                processing_logger.info("🔍 Analyzing template cells for format detection...")
                
                # ULTRA-FAST OPTIMIZATION: Only process cells with meaningful formats
                # This reduces processing from 36k+ cells to ~100-500 cells
                cell_count = 0
                processed_cells = 0
                
                for row in template_ws.iter_rows():
                    for tcell in row:
                        coord = tcell.coordinate
                        template_coords.add(coord)
                        cell_count += 1
                        
                        # FLEXIBLE FILTER: Process all cells with values or meaningful content
                        fmt = getattr(tcell, 'number_format', None)
                        has_value = tcell.value is not None and tcell.value != ''
                        has_special_format = fmt and fmt not in ['General', '@', '0', '0.00']
                        has_formula = hasattr(tcell, 'data_type') and tcell.data_type == 'f'
                        
                        # Process cells with values, special formatting, OR formulas (including totals)
                        if has_value or has_special_format or has_formula:
                            processed_cells += 1
                            
                            # Enhanced format detection with comprehensive analysis
                            format_info = {
                                'is_percentage': False,
                                'is_currency': False,
                                'is_number': False,
                                'is_date': False,
                                'number_format': str(fmt) if fmt else None,
                                'has_formula': False,
                                'format_confidence': 1.0,
                                'consolidation_method': 'sum'  # Default to sum
                            }
                            
                            try:
                                # Enhanced format detection using new helper methods
                                if fmt:
                                    fmt_str = str(fmt)
                                    
                                    if self._is_percentage_format(fmt_str):
                                        format_info['is_percentage'] = True
                                        format_info['consolidation_method'] = 'average'
                                        processing_logger.info(f"📊 Percentage cell detected: {coord} with format: {fmt}")
                                    
                                    elif self._is_currency_format(fmt_str):
                                        format_info['is_currency'] = True
                                        format_info['consolidation_method'] = 'sum'
                                    
                                    elif self._is_number_format(fmt_str):
                                        format_info['is_number'] = True
                                        format_info['consolidation_method'] = 'sum'
                                    
                                    elif self._is_date_format(fmt_str):
                                        format_info['is_date'] = True
                                        format_info['consolidation_method'] = 'sum'
                                
                                # Formula detection (simplified)
                                if hasattr(tcell, 'data_type') and tcell.data_type == 'f':
                                    format_info['has_formula'] = True
                                elif isinstance(tcell.value, str) and str(tcell.value).startswith('='):
                                    format_info['has_formula'] = True
                                
                                coord_format_info[coord] = format_info
                                
                            except Exception:
                                # Silent error handling to avoid logging overhead
                                continue
                
                processing_logger.info(f"📊 Enhanced format detection completed. Processed {processed_cells} relevant cells out of {cell_count} total cells")
                
                # Log format summary for debugging
                percent_count = len([c for c in coord_format_info.values() if c.get('is_percentage')])
                currency_count = len([c for c in coord_format_info.values() if c.get('is_currency')])
                number_count = len([c for c in coord_format_info.values() if c.get('is_number')])
                date_count = len([c for c in coord_format_info.values() if c.get('is_date')])
                
                processing_logger.info(f"📊 Format summary: {percent_count} percentage cells, {currency_count} currency cells, {number_count} number cells, {date_count} date cells")
                
                # Log all percentage cells for debugging
                percent_cells = [coord for coord, info in coord_format_info.items() if info.get('is_percentage')]
                if percent_cells:
                    processing_logger.info(f"📊 Percentage cells found: {percent_cells[:10]}{'...' if len(percent_cells) > 10 else ''}")

                # Propagate number format across merged ranges with enhanced format inheritance
                try:
                    for mrange in getattr(template_ws, 'merged_cells', []).ranges:
                        try:
                            min_row = mrange.min_row
                            max_row = mrange.max_row
                            min_col = mrange.min_col
                            max_col = mrange.max_col
                            master_cell = template_ws.cell(row=min_row, column=min_col)
                            master_coord = master_cell.coordinate
                            
                            # Get master cell format info
                            master_format = coord_format_info.get(master_coord, {})
                            
                            # Propagate format to all cells in merged range
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    coord = template_ws.cell(row=r, column=c).coordinate
                                    template_coords.add(coord)
                                    # Inherit master cell format
                                    coord_format_info[coord] = master_format.copy()
                        except Exception:
                            continue
                except Exception:
                    pass
                    
            except Exception:
                template_ws = None
                coord_format_info = {}
                template_coords = None
            # PERFORMANCE FIX: Skip the slow format update process
            # The format update was taking 16+ minutes per file and is not necessary
            # for percentage averaging to work correctly
            processing_logger.info(f"🔧 Format info available: {len(coord_format_info)} coordinates")
            
            # Log percentage cells found for debugging
            percent_cells = [coord for coord, info in coord_format_info.items() if info.get('is_percentage')]
            if percent_cells:
                processing_logger.info(f"📊 Percentage cells detected: {percent_cells[:5]}{'...' if len(percent_cells) > 5 else ''}")
            
            # PROFESSIONAL APPROACH: NEVER modify source files (too slow!)
            # Instead, handle format conversion ON-THE-FLY during reading
            enable_format_standardization = self.settings.get('data_processing', {}).get('enable_format_standardization', False)
            
            self.progress.emit(2)
            
            if enable_format_standardization:
                processing_logger.warning("⚠️ Format standardization ENABLED - this will be VERY SLOW!")
                processing_logger.warning("⚠️ Modifying and saving source files takes ~30-60 seconds per file")
                processing_logger.warning("⚠️ For 100 files, this could take 1-2 HOURS!")
                processing_logger.info("💡 TIP: Disable this setting for 20x faster processing")
                self._update_submitted_files_format(files, coord_format_info)
            else:
                processing_logger.info("⚡ Format standardization DISABLED - using FAST on-the-fly conversion")
                processing_logger.info("⚡ Source files will NOT be modified (read-only mode)")
                processing_logger.info("⚡ Format conversion happens during reading (in-memory only)")
            
            self.progress.emit(4)
            
            # OPTIMIZATION: Process files in batches for better performance
            total_files = len(files)
            processing_logger.info(f"📁 Processing {total_files} files...")
            
            if enable_format_standardization:
                processing_logger.info("✅ Files standardized to template format - ready for consolidation")
            else:
                processing_logger.info("⚡ ULTRA-FAST mode - processing files as-is")
            
            for idx, file in enumerate(files, 1):
                try:
                    # Progress indicator for large batches
                    if total_files > 10 and idx % max(1, total_files // 10) == 0:
                        processing_logger.info(f"📊 Progress: {idx}/{total_files} files processed ({idx/total_files*100:.1f}%)")
                    
                    ext = os.path.splitext(file)[1].lower()
                    file_label = os.path.splitext(os.path.basename(file))[0]

                    if ext in ('.xlsx', '.xls'):
                        wb = openpyxl.load_workbook(file, data_only=True, read_only=bool(self.settings.get('performance', {}).get('memory_optimization')))
                        ws = self._get_worksheet(wb, "source")
                        if validate_structure and template_ws is not None:
                            try:
                                if (ws.max_row != template_ws.max_row) or (ws.max_column != template_ws.max_column):
                                    if stop_on_error:
                                        wb.close()
                                        filename = os.path.basename(file)
                                        error_msg = (f"File Structure Mismatch\n\n"
                                                   f"File '{filename}' has a different structure than the template:\n\n"
                                                   f"Template: {template_ws.max_row} rows × {template_ws.max_column} columns\n"
                                                   f"File: {ws.max_row} rows × {ws.max_column} columns\n\n"
                                                   f"Solution: Ensure all files have the same structure as the template, "
                                                   f"or disable structure validation in settings.")
                                        self.finished.emit("error", error_msg)
                                        return
                                    wb.close()
                                    self.file_processed.emit(os.path.basename(file))
                                    continue
                            except Exception:
                                pass
                        if load_cells is not None:
                            cells_iter = load_cells(ws, self.settings)
                        else:
                            cells_iter = (cell for row in ws.iter_rows() for cell in row)

                        for cell in cells_iter:
                            value = cell.value
                            coord = cell.coordinate
                            
                            # If template is present, skip cells not defined in template
                            if template_coords is not None and coord not in template_coords:
                                continue
                            
                            # Get template format information for this coordinate
                            format_info = coord_format_info.get(coord, {})
                            
                            # CRITICAL FIX: Skip formulas in SOURCE files to prevent double-counting
                            # Check if THIS cell (in source file) has a formula, not template
                            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                # This cell in SOURCE file is a formula - SKIP IT!
                                # Formulas often reference other cells being consolidated, causing double-counting
                                processing_logger.debug(f"⏩ Skipping formula cell {coord} in {file_label} to prevent double-counting")
                                continue
                            
                            # Skip empty cells
                            if value is None or value == '':
                                continue
                            
                            # FLEXIBLE: Handle total cells
                            include_totals = self.settings.get('data_processing', {}).get('include_totals', True)
                            if not include_totals and self._is_total_cell(cell):
                                continue
                            
                            # Process the cell value
                            val = self._process_cell_value_with_format_verification(
                                value, format_info, coord, file_label, wb, stop_on_error
                            )
                            if val is None:
                                continue
                                
                            # Validate value against settings
                            if validate_value is not None and not validate_value(val, self.settings):
                                if stop_on_error:
                                    wb.close()
                                    filename = os.path.basename(file)
                                    error_msg = (f"Data Validation Error\n\n"
                                               f"Value {val} at cell {cell.coordinate} in file '{filename}' "
                                               f"is outside the allowed range.\n\n"
                                               f"Please check the data in this file or adjust the validation settings.")
                                    self.finished.emit("error", error_msg)
                                    return
                                continue
                            
                            # Enhanced processing based on template format requirements
                            consolidation_method = format_info.get('consolidation_method', 'sum')
                            
                            if consolidation_method == 'average':
                                # For percentage cells: accumulate for average calculation
                                # Count behavior depends on exclude_zero_percent setting
                                current_total = totals.get(coord)
                                totals[coord] = (current_total + val) if current_total is not None else val
                                
                                # Initialize count to total files on first encounter
                                if coord not in percent_counts:
                                    if self.exclude_zero_percent:
                                        # When excluding zeros: only count files with non-zero values
                                        percent_counts[coord] = 0
                                    else:
                                        # Default: count all files (including files with 0% values)
                                        percent_counts[coord] = total_files_count
                                
                                # If excluding zeros, increment count only for non-zero values
                                if self.exclude_zero_percent and val != 0:
                                    percent_counts[coord] += 1
                                
                                # Enhanced debug logging for percentage cells
                                count_mode = "non-zero files" if self.exclude_zero_percent else "all files"
                                processing_logger.info(f"📊 Percentage cell {coord}: {val} (from {file_label}) - Total: {totals[coord]}, Count: {percent_counts[coord]} ({count_mode})")
                                
                            else:
                                # For currency, number, and unformatted cells: sum values
                                # Zero values don't affect sum, but are included conceptually
                                current_total = totals.get(coord)
                                totals[coord] = (current_total + val) if current_total is not None else val
                                
                                # Enhanced debug logging for all cell types
                                cell_type = "currency" if format_info.get('is_currency') else "number" if format_info.get('is_number') else "unformatted"
                                processing_logger.info(f"🔢 {cell_type.title()} cell {coord}: {val} (from {file_label}) - Total: {totals[coord]}")
                            
                            # Track contributions for detailed reporting
                            if coord not in contributions:
                                contributions[coord] = {}
                            prev = contributions[coord].get(file_label)
                            contributions[coord][file_label] = (prev + val) if prev is not None else val
                        wb.close()
                    elif ext == '.csv':
                        self.file_processed.emit(os.path.basename(file))
                        continue
                    self.file_processed.emit(os.path.basename(file))
                except Exception as e:
                    error_msg = self._get_file_error_message(file, e)
                    # Log the detailed error for debugging
                    print(f"File processing failed: {error_msg}")
                    try:
                        from src.modules.google_sheets_reporter import GoogleSheetsErrorReporter
                        error_reporter = GoogleSheetsErrorReporter("1.0.1")
                        error_reporter.report_error(
                            type(e), e, e.__traceback__,
                            triggered_by="File Processing in ConsolidationWorker",
                            user_file=file
                        )
                    except Exception:
                        pass

                self.progress.emit(5 + int(idx / max(len(files), 1) * 80))

            from openpyxl.comments import Comment
            from openpyxl.styles import Border, Side
            thin_orange = Border(
                left=Side(style='thin', color='FF8C00'),
                right=Side(style='thin', color='FF8C00'),
                top=Side(style='thin', color='FF8C00'),
                bottom=Side(style='thin', color='FF8C00')
            )

            # Determine whether we should overwrite formulas in the output template
            overwrite_output_formulas = (
                self.settings.get('output_handling', {}).get('overwrite_output_formulas', True)
            )

            for coord, value in totals.items():
                cell = output_ws[coord]
                if isinstance(cell, MergedCell):
                    continue
                # Optionally overwrite formulas in the template/output to ensure accurate consolidated totals
                if not overwrite_output_formulas:
                    try:
                        if getattr(cell, 'data_type', None) == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                            continue
                    except Exception:
                        pass

                # Enhanced consolidation logic with format-aware processing
                format_info = coord_format_info.get(coord, {})
                consolidation_method = format_info.get('consolidation_method', 'sum')
                
                # Enhanced debugging for format detection
                processing_logger.info(f"🔍 Consolidating {coord}: Format info = {format_info}")
                processing_logger.info(f"🔍 Consolidation method: {consolidation_method}, Value: {value}")
                
                try:
                    if consolidation_method == 'average':
                        # For percentage cells: calculate average (total ÷ count) and format as percentage
                        count = max(1, percent_counts.get(coord, 1))
                        avg_value = float(value / Decimal(count))
                        
                        # Enhanced debug logging for final consolidation
                        processing_logger.info(f"🎯 Final percentage consolidation for {coord}: Total={value}, Count={count}, Average={avg_value} ({avg_value:.2f}%)")
                        
                        # Set the calculated average value - values are already in percentage points, just convert to decimal for Excel
                        # Excel expects percentage values as decimals (e.g., 0.825 for 82.5%)
                        cell.value = avg_value / 100
                        
                        # Ensure the cell maintains percentage format from template
                        template_format = format_info.get('number_format', '0.00%')
                        cell.number_format = template_format
                        
                        processing_logger.info(f"✅ {coord}: Set to {avg_value/100} ({avg_value:.2f}%) with format {template_format}")
                        
                    else:
                        # For currency, number, and unformatted cells: sum values
                        cell.value = float(value)
                        
                        # Apply appropriate formatting based on cell type
                        if format_info.get('is_currency', False):
                            template_format = format_info.get('number_format', '$#,##0.00')
                            cell.number_format = template_format
                            processing_logger.info(f"✅ {coord}: Currency sum = {float(value)} with format {template_format}")
                            
                        elif format_info.get('is_number', False):
                            template_format = format_info.get('number_format', '#,##0.00')
                            cell.number_format = template_format
                            processing_logger.info(f"✅ {coord}: Number sum = {float(value)} with format {template_format}")
                            
                        else:
                            # Default: sum values without special formatting
                            processing_logger.info(f"✅ {coord}: Unformatted sum = {float(value)}")
                        
                except Exception:
                    # Fallback to basic value assignment
                    cell.value = float(value) if value is not None else 0
                file_map = contributions.get(coord, {})
                if file_map:
                    items = sorted(file_map.items(), key=lambda x: x[0].lower())
                    max_name = max((len(n) for n, _ in items), default=4)
                    header = "Consolidation Summary\n"
                    header += f"Cell: {coord}\n"
                    
                    # Enhanced summary based on cell format
                    format_info = coord_format_info.get(coord, {})
                    is_percent = format_info.get('is_percentage', False)
                    
                    if is_percent:
                        count = max(1, int(percent_counts.get(coord, 1)))
                        avg_val = (value / Decimal(count))
                        # BUG FIX: avg_val is already in percentage points, don't multiply by 100!
                        num_contributors = len([v for v in file_map.values() if v != 0])
                        
                        if self.exclude_zero_percent:
                            # Excluding zeros: count only includes files with non-zero values
                            header += f"Average: {float(avg_val):,.2f}% (from {count} files with values"
                            if num_contributors != count:
                                header += f", {num_contributors} non-zero"
                            header += ", zero values excluded)\n\n"
                        else:
                            # Default: count includes ALL files (missing/empty cells treated as 0%)
                            header += f"Average: {float(avg_val):,.2f}% (from {count} files"
                            if num_contributors < count:
                                header += f", {num_contributors} with values, {count - num_contributors} empty"
                            header += ")\n\n"
                    elif format_info.get('is_currency', False):
                        header += f"Total: ${float(value):,.2f}\n\n"
                    elif format_info.get('is_number', False):
                        header += f"Total: {float(value):,.2f}\n\n"
                    else:
                        header += f"Total: {float(value):,.2f}\n\n"
                    header += "Contributors (file  |  value)\n"
                    header += "-" * (max(26, max_name + 10)) + "\n"
                    lines = []
                    for name, v in items:
                        pad = " " * (max_name - len(name))
                        try:
                            format_info = coord_format_info.get(coord, {})
                            if format_info.get('is_percentage', False):
                                # BUG FIX: v is already in percentage points, don't multiply by 100!
                                lines.append(f"{name}{pad}  |  {float(v):,.2f}%")
                            elif format_info.get('is_currency', False):
                                lines.append(f"{name}{pad}  |  ${float(v):,.2f}")
                            else:
                                lines.append(f"{name}{pad}  |  {float(v):,.2f}")
                        except Exception:
                            lines.append(f"{name}{pad}  |  {v}")
                    body = "\n".join(lines)
                    comment_text = header + body
                    max_len = 32000
                    if len(comment_text) > max_len:
                        comment_text = comment_text[: max_len - 21] + "\n... (truncated)"
                    comment = Comment(comment_text, "Excel Consolidator")
                    comment.width = min(520, 200 + max_name * 7)
                    comment.height = min(600, 140 + len(items) * 14)
                    cell.comment = comment
                    cell.border = thin_orange

            self.progress.emit(90)
            os.makedirs(self.save_folder, exist_ok=True)
            date_str = datetime.now().strftime("%b %d %Y")
            output_name = f"Consolidated - {date_str}.xlsm" if keep_vba else f"Consolidated - {date_str}.xlsx"
            output_path = os.path.join(self.save_folder, output_name)
            
            try:
                contrib_ws = output_wb.create_sheet("Contributions")
                contrib_ws["A1"] = "CONTRIBUTIONS INDEX"
                contrib_ws["A1"].font = Font(bold=True, size=14, color="2F5597")
                contrib_ws.merge_cells('A1:D1')
                contrib_ws["A3"] = "Search (use column filters):"
                contrib_ws["A5"] = "Cell"
                contrib_ws["B5"] = "File Name"
                contrib_ws["C5"] = "Contribution"
                r = 6
                coord_to_first_row = {}
                # Sort coordinates in natural Excel order (A1, A2, ..., B1, ...)
                def _col_to_num(col_letters):
                    result = 0
                    for ch in col_letters:
                        if 'a' <= ch <= 'z':
                            ch = chr(ord(ch) - 32)
                        result = result * 26 + (ord(ch) - 64)
                    return result
                def _cell_sort_key(cell_ref):
                    col = ""
                    row_str = ""
                    for ch in cell_ref:
                        if ch.isalpha():
                            col += ch
                        elif ch.isdigit():
                            row_str += ch
                    return (_col_to_num(col), int(row_str) if row_str else 0)

                for coord in sorted(contributions.keys(), key=_cell_sort_key):
                    file_map = contributions.get(coord, {})
                    # Iterate through the complete set of files; fill 0 where missing
                    for fname in all_file_labels:
                        v = file_map.get(fname, 0)
                        contrib_ws[f"A{r}"] = coord
                        contrib_ws[f"B{r}"] = fname
                        try:
                            format_info = coord_format_info.get(coord, {})
                            v_out = v
                            
                            if format_info.get('is_percentage', False):
                                # PERCENTAGE VALUES - Accurate Display in Contributions Sheet
                                # v_out is in percentage points (e.g., 84.36 for 84.36%)
                                # Excel needs decimal format (0.8436) with percentage format to display as 84.36%
                                contrib_ws[f"C{r}"] = float(v_out) / 100  # Convert to Excel decimal
                                template_format = format_info.get('number_format', '0.00%')
                                contrib_ws[f"C{r}"].number_format = template_format
                                # Result: Displays as 84.36% (CORRECT)
                                
                            elif format_info.get('is_currency', False):
                                # CURRENCY VALUES - Accurate Display
                                # v_out is already in correct format (e.g., 1234.56)
                                contrib_ws[f"C{r}"] = float(v_out)
                                template_format = format_info.get('number_format', '$#,##0.00')
                                contrib_ws[f"C{r}"].number_format = template_format
                                # Result: Displays as $1,234.56 (CORRECT)
                                
                            elif format_info.get('is_number', False):
                                # NUMBER VALUES - Accurate Display
                                # v_out is already in correct format (e.g., 1234.56)
                                contrib_ws[f"C{r}"] = float(v_out)
                                template_format = format_info.get('number_format', '#,##0.00')
                                contrib_ws[f"C{r}"].number_format = template_format
                                # Result: Displays as 1,234.56 (CORRECT)
                                
                            else:
                                # DEFAULT - Unformatted values
                                contrib_ws[f"C{r}"] = float(v_out)
                                
                        except Exception:
                            # Fallback: Use raw value if formatting fails
                            contrib_ws[f"C{r}"] = v
                        if coord not in coord_to_first_row:
                            coord_to_first_row[coord] = r
                        r += 1
                    # Add a visual break between groups of the same cell reference
                    # This blank row helps users identify each group easily
                    r += 1
                if r > 6:
                    contrib_ws.auto_filter.ref = f"A5:C{r-1}"
                contrib_ws.column_dimensions['A'].width = 12
                contrib_ws.column_dimensions['B'].width = 40
                contrib_ws.column_dimensions['C'].width = 16
                try:
                    for coord in totals.keys():
                        first_row = coord_to_first_row.get(coord)
                        if first_row:
                            cell = output_ws[coord]
                            if isinstance(cell, MergedCell):
                                continue
                            link = f"#'Contributions'!A{first_row}"
                            cell.hyperlink = link
                except Exception:
                    pass
                # Create a plain consolidated sheet with full formatting (but no hyperlinks/comments)
                try:
                    plain_ws = output_wb.create_sheet("Consolidated (Plain)")
                    
                    # Copy all formatting from the main output sheet
                    # This includes: fonts, fills, borders, alignments, number formats, column widths, row heights, merged cells
                    
                    # Copy merged cells first
                    for merged_range in output_ws.merged_cells.ranges:
                        plain_ws.merge_cells(str(merged_range))
                    
                    # Copy column widths
                    for col_letter, col_dim in output_ws.column_dimensions.items():
                        plain_ws.column_dimensions[col_letter].width = col_dim.width
                    
                    # Copy row heights
                    for row_num, row_dim in output_ws.row_dimensions.items():
                        plain_ws.row_dimensions[row_num].height = row_dim.height
                    
                    # Copy all cells with their formatting and values
                    for row in output_ws.iter_rows():
                        for cell in row:
                            plain_cell = plain_ws[cell.coordinate]
                            
                            # Skip merged cells (already handled above)
                            if isinstance(cell, MergedCell) or isinstance(plain_cell, MergedCell):
                                continue
                            
                            # Copy value (but not formula - use the calculated value)
                            plain_cell.value = cell.value
                            
                            # Copy all formatting
                            if cell.has_style:
                                plain_cell.font = copy(cell.font)
                                plain_cell.border = copy(cell.border)
                                plain_cell.fill = copy(cell.fill)
                                plain_cell.number_format = copy(cell.number_format)
                                plain_cell.protection = copy(cell.protection)
                                plain_cell.alignment = copy(cell.alignment)
                            
                            # Explicitly do NOT copy hyperlinks or comments
                            # (plain_cell.hyperlink and plain_cell.comment remain None)
                            
                except Exception:
                    pass
            except Exception:
                pass

            # Remove "Sheet 2" if it exists (cleanup unwanted default sheets)
            try:
                if "Sheet 2" in output_wb.sheetnames:
                    sheet_to_remove = output_wb["Sheet 2"]
                    output_wb.remove(sheet_to_remove)
            except Exception:
                pass

            backup_target = None
            if ensure_backup is not None:
                backup_target = ensure_backup(self.save_folder, self.settings, os.path.basename(output_path))
            try:
                output_wb.save(output_path)
            except Exception as e:
                error_msg = self._get_save_error_message(e, output_path)
                self.finished.emit("error", error_msg)
                return
            if backup_target:
                try:
                    import shutil
                    shutil.copy2(output_path, backup_target)
                except Exception:
                    pass
            self.progress.emit(100)
            self.finished.emit("success", output_path)
        except Exception as e:
            error_message = self._get_user_friendly_error_message(e)
            self.finished.emit("error", error_message)

def _num_to_col_name(n):
    """Convert 1-based index to Excel column name (1->A, 27->AA)."""
    name = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name

# ---------------- Main App ----------------
class ExcelProcessorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION}")
        self.setGeometry(400, 200, 900, 650)
        icon_path = self.resource_path("assets/icons/app.ico")
        self.setWindowIcon(QIcon(icon_path))

        self.template_path = None
        self.excel_folder = None
        self.save_folder = None
        self.advanced_settings = None

        self.error_reporter = None
        # Auto-update dialog guard removed
        self.setup_error_reporting_and_updates()

        self.apply_modern_style()
        self.initUI()

    def resource_path(self, relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    
    def setup_error_reporting_and_updates(self):
        try:
            if ERROR_REPORTING_ENABLED:
                self.error_reporter, _ = setup_google_sheets_error_reporting(APP_VERSION)
                if self.error_reporter:
                    print("Google Sheets error reporting system initialized successfully")
                else:
                    print("Warning: Google Sheets error reporting system failed to initialize")
            # Auto-update feature has been removed
        except Exception as e:
            print(f"Warning: Failed to setup error reporting system: {e}")
    
    
    

    def closeEvent(self, event):
        # Auto-update cleanup removed
        event.accept()
    
    def report_error_with_context(self, exc_type, exc_value, exc_traceback, triggered_by="Unknown", user_file=None):
        try:
            if self.error_reporter:
                self.error_reporter.report_error(exc_type, exc_value, exc_traceback, triggered_by, user_file)
                QMessageBox.information(
                    self,
                    "Error Reported",
                    self.error_reporter.get_user_friendly_message()
                )
        except Exception as e:
            print(f"Error reporting failed: {e}")

    def apply_modern_style(self):
        check_svg = self.resource_path("assets/icons/check.svg").replace("\\", "/")
        check_disabled_svg = self.resource_path("assets/icons/check_disabled.svg").replace("\\", "/")
        css = """
            /* Base */
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
                color: #1f2937; /* slate-800 */
            }
            QMainWindow, QWidget, QDialog {
                background-color: #deeaee; /* requested background */
            }
            /* Dialog text widgets should inherit group box white background */
            QDialog QLabel, QDialog QCheckBox, QDialog QRadioButton {
                background: transparent;
            }
            /* Read-only rich text areas in dialogs should be white */
            QDialog QTextEdit[readOnly="true"] {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }

            /* Headings */
            QLabel#HeaderTitle { color: #0f766e; } /* teal-700 */
            QLabel#HeaderSubtitle { color: #6b7280; } /* gray-500 */

            /* Buttons - default */
            QPushButton {
                border: 1px solid #e5e7eb; /* gray-200 */
                background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #f5f7fb);
                color: #111827; /* gray-900 */
                border-radius: 8px;
                padding: 10px 14px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #eef2ff; /* indigo-50 */
                border-color: #c7d2fe; /* indigo-200 */
            }
            QPushButton:pressed {
                background-color: #e0e7ff; /* indigo-100 */
            }
            QPushButton:disabled {
                background-color: #f3f4f6;
                color: #9ca3af;
                border-color: #e5e7eb;
            }

            /* Group boxes */
            QGroupBox {
                font-weight: 600;
                border: 1px solid #e5e7eb;
                border-radius: 10px;
                margin-top: 12px;
                padding-top: 22px; /* lower the floating title slightly */
                background: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 2px 6px; /* add slight top padding to lower text */
                color: #0f766e; /* match title color */
                background: transparent; /* transparent label background */
            }

            /* Inputs */
            QLineEdit, QComboBox, QTextEdit, QSpinBox, QDoubleSpinBox {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                padding: 8px 10px;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus,
            QSpinBox:focus, QDoubleSpinBox:focus {
                border: 1px solid #34d399; /* emerald-400 */
                outline: none;
            }

            /* Spin boxes: align button area with rounded input */
            QAbstractSpinBox {
                padding-right: 30px; /* space for buttons */
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }
            QAbstractSpinBox::up-button,
            QAbstractSpinBox::down-button {
                subcontrol-origin: border;
                width: 28px;
                background: #ffffff;
                border-left: 1px solid #e5e7eb;
                margin: 0;
            }
            QAbstractSpinBox::up-button {
                subcontrol-position: right top;
                border-top-right-radius: 8px;
            }
            QAbstractSpinBox::down-button {
                subcontrol-position: right bottom;
                border-bottom-right-radius: 8px;
            }
            QAbstractSpinBox::up-button:hover,
            QAbstractSpinBox::down-button:hover { background: #f9fafb; }

            /* Tabs */
            QTabWidget::pane { border: 1px solid #e5e7eb; border-radius: 10px; }
            QTabBar::tab {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                padding: 8px 14px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 4px;
            }
            QTabBar::tab:selected { color: #0f766e; border-bottom: 2px solid #10b981; }
            QTabBar::tab:hover { background: #f8fafc; }

            /* Lists */
            QListWidget {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }

            /* Progress */
            QProgressBar {
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                text-align: center;
                background-color: #ffffff;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #10b981; /* emerald-500 */
                border-radius: 8px;
            }

            /* Splitter */
            QSplitter { border: none; }
            QSplitter::handle { background: #edf2f7; width: 6px; }
            QSplitter::handle:hover { background: #e2e8f0; }

            /* Tooltips */
            QToolTip {
                background-color: #111827; /* gray-900 */
                color: #f9fafb; /* gray-50 */
                border: 1px solid #111827;
                padding: 6px 8px;
                border-radius: 6px;
                font-size: 12px;
            }

            /* Checkboxes and Radios */
            QCheckBox::indicator {
                width: 12px; height: 12px;
                border-radius: 6px;
                border: 2px solid #15803d; /* green-700 */
                background: #ffffff; /* white box */
            }
            QCheckBox::indicator:hover { border-color: #16a34a; }
            QCheckBox::indicator:checked {
                background: #ffffff; /* keep white like reference */
                border: 2px solid #15803d;
                image: url(./resources/check.svg);
            }
            QCheckBox::indicator:disabled {
                border-color: #e5e7eb;
                background: #f3f4f6;
                image: url(./resources/check_disabled.svg);
            }

            QRadioButton::indicator {
                width: 16px; height: 16px;
                border-radius: 8px;
                border: 2px solid #d1d5db;
                background: #ffffff;
            }
            QRadioButton::indicator:checked {
                border: 6px solid #16a34a; /* inner dot via border */
            }

            /* Button variants via objectName */
            QPushButton#PrimaryButton { background: #0ea5e9; color: white; border: 1px solid #0ea5e9; }
            QPushButton#PrimaryButton:hover { background: #0284c7; border-color: #0284c7; }
            QPushButton#SuccessButton { background: #10b981; color: white; border: 1px solid #10b981; }
            QPushButton#SuccessButton:hover { background: #059669; border-color: #059669; }
            QPushButton#DangerButton { background: #ef4444; color: white; border: 1px solid #ef4444; }
            QPushButton#DangerButton:hover { background: #dc2626; border-color: #dc2626; }
            QPushButton#TertiaryButton { background: #ffffff; color: #111827; border: 1px solid #e5e7eb; }
            QPushButton#TertiaryButton:hover { background: #f9fafb; }
        """
        css = css.replace("./resources/check.svg", check_svg)
        css = css.replace("./resources/check_disabled.svg", check_disabled_svg)
        self.setStyleSheet(css)

    def initUI(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Update indicator removed
        
        watermark = QLabel("© 2025 Izak. All rights reserved.", self)
        watermark.setStyleSheet("""
            QLabel {
                color: rgba(0, 0, 0, 0.2);
                font-size: 12px;
                font-style: italic;
            }
        """)
        watermark.setAlignment(Qt.AlignRight | Qt.AlignBottom)
        watermark.setGeometry(0, 0, 200, 30)
        
        def updateWatermarkPos():
            watermark.move(self.width() - watermark.width() - 20, 
                         self.height() - watermark.height() - 5)
        
        self.resizeEvent = lambda e: updateWatermarkPos()

        header_frame = QFrame()
        header_frame.setObjectName("AppHeader")
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(20, 14, 20, 12)
        header_layout.setSpacing(16)
                
        logo_label = QLabel()
        logo_path = self.resource_path("assets/icons/logo.png")
        logo_pixmap = QPixmap(logo_path)
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
        scaled_pixmap = logo_pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        logo_label.setPixmap(scaled_pixmap)
        logo_label.setFixedSize(100, 100)
        logo_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(logo_label)
        
        title_container = QVBoxLayout()
        title_container.setContentsMargins(0, 0, 0, 0)
        title_container.setSpacing(2)
        title = QLabel("Excel Consolidator")
        title.setObjectName("HeaderTitle")
        title.setFont(QFont("Segoe UI", 22, QFont.Bold))
        title.setStyleSheet("color: #0f766e;")
        subtitle = QLabel("Sum values from multiple Excel files while preserving formatting")
        subtitle.setObjectName("HeaderSubtitle")
        subtitle.setFont(QFont("Segoe UI", 10))
        subtitle.setStyleSheet("color: #6b7280;")
        title_container.addWidget(title)
        title_container.addWidget(subtitle)
        header_layout.addLayout(title_container)
        header_layout.setAlignment(title_container, Qt.AlignVCenter | Qt.AlignLeft)
        
        header_layout.addStretch()
        main_layout.addWidget(header_frame)
        header_sep = QFrame()
        header_sep.setObjectName("Separator")
        header_sep.setFixedHeight(1)
        main_layout.addWidget(header_sep)

        content_splitter = QSplitter(Qt.Horizontal)
        
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(8, 4, 8, 8)
        left_layout.setSpacing(12)
        
        template_group = QGroupBox("Step 1: Select Template File")
        template_group.setStyleSheet("QGroupBox::title{background:transparent;color:#0f766e;}")
        template_layout = QVBoxLayout()
        template_layout.setContentsMargins(12, 12, 12, 12)
        template_layout.setSpacing(10)
        self.template_btn = QPushButton("📄 Browse Template File")
        self.template_btn.setObjectName("TertiaryButton")
        self.template_btn.clicked.connect(self.open_template_file)
        self.template_label = QLabel("No template selected")
        self.template_label.setWordWrap(True)
        self.template_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
        template_layout.addWidget(self.template_btn)
        template_layout.addWidget(self.template_label)
        template_group.setLayout(template_layout)
        left_layout.addWidget(template_group)

        folder_group = QGroupBox("Step 2: Select Excel Folder")
        folder_group.setStyleSheet("QGroupBox::title{background:transparent;color:#0f766e;}")
        folder_layout = QVBoxLayout()
        folder_layout.setContentsMargins(12, 12, 12, 12)
        folder_layout.setSpacing(10)
        self.folder_btn = QPushButton("📂 Browse Excel Folder")
        self._style_button(self.folder_btn, "#388E3C")
        self.folder_btn.clicked.connect(self.open_excel_folder)
        self.folder_label = QLabel("No folder selected")
        self.folder_label.setWordWrap(True)
        self.folder_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
        folder_layout.addWidget(self.folder_btn)
        folder_layout.addWidget(self.folder_label)
        
        # Checkbox to exclude zero values from percentage averages
        self.exclude_zero_percent = QCheckBox("Exclude zero (0%) values when calculating percentage averages")
        self.exclude_zero_percent.setChecked(False)  # Default: include all values
        self.exclude_zero_percent.setStyleSheet("""
            QCheckBox {
                color: #374151;
                font-size: 12px;
                padding: 4px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
            }
        """)
        self.exclude_zero_percent.setToolTip(
            "When DISABLED (default): All files are included in percentage averages, including files with 0% values.\n"
            "Example: 100%, 50%, 0% → Average = (100 + 50 + 0) ÷ 3 = 50%\n\n"
            "When ENABLED: Files with 0% values are excluded from the average calculation.\n"
            "Example: 100%, 50%, 0% → Average = (100 + 50) ÷ 2 = 75%\n\n"
            "This only affects percentage-formatted cells. Other cell types are unaffected."
        )
        folder_layout.addWidget(self.exclude_zero_percent)
        
        folder_group.setLayout(folder_layout)
        left_layout.addWidget(folder_group)

        save_group = QGroupBox("Step 3: Select Save Location")
        save_group.setStyleSheet("QGroupBox::title{background:transparent;color:#0f766e;}")
        save_layout = QVBoxLayout()
        save_layout.setContentsMargins(12, 12, 12, 12)
        save_layout.setSpacing(10)
        self.save_btn = QPushButton("💾 Browse Save Location")
        self._style_button(self.save_btn, "#D32F2F")
        self.save_btn.clicked.connect(self.open_save_folder)
        self.save_label = QLabel("No save location selected")
        self.save_label.setWordWrap(True)
        self.save_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
        save_layout.addWidget(self.save_btn)
        save_layout.addWidget(self.save_label)
        save_group.setLayout(save_layout)
        left_layout.addWidget(save_group)

        settings_run_layout = QVBoxLayout()
        self.advanced_btn = QPushButton("⚙️ Advanced Settings")
        self.advanced_btn.setObjectName("TertiaryButton")
        self.advanced_btn.clicked.connect(self.open_advanced_settings)
        self.advanced_btn.setEnabled(False)  # Disabled - feature not fully functional
        self.advanced_btn.setVisible(False)  # Hidden to prevent accidental usage
        settings_run_layout.addWidget(self.advanced_btn)
        self.run_btn = QPushButton("🚀 Run Consolidation")
        self.run_btn.setObjectName("SuccessButton")
        self.run_btn.clicked.connect(self.run_processing)
        settings_run_layout.addWidget(self.run_btn)
        left_layout.addLayout(settings_run_layout)
        left_layout.addStretch()
        content_splitter.addWidget(left_widget)
        
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(8, 4, 8, 8)
        right_layout.setSpacing(12)
        
        info_group = QGroupBox("Information")
        info_group.setStyleSheet("QGroupBox::title{background:transparent;color:#0f766e;}")
        info_layout = QVBoxLayout()
        info_layout.setContentsMargins(12, 12, 12, 12)
        info_layout.setSpacing(10)
        info_text = QTextEdit()
        info_text.setReadOnly(True)
        info_text.setHtml("""
            <h3>About Excel Consolidator</h3>
            <p>This tool helps you sum values from multiple Excel files while preserving all formatting from your template.</p>
            <h4>Core Features:</h4>
            <ul>
                <li>Sums numeric values from multiple Excel files</li>
                <li>Preserves colors, fonts, borders, and cell formatting</li>
                <li>Maintains column widths and row heights</li>
                <li>Handles merged cells correctly</li>
                <li>Shows real-time progress</li>
            </ul>
            <h4>Output Features:</h4>
            <ul>
                <li><b>Multiple sheets:</b> Main, Contributions, and Plain versions</li>
                <li><b>Consolidated (Plain):</b> Fully formatted data without hyperlinks</li>
                <li>Multi-format support (XLSX, XLS)</li>
                <li>Smart data validation & conversion</li>
                <li>Duplicate file detection & handling</li>
                <li>Automatic backup & recovery</li>
            </ul>
            <h4>Interactive Verification:</h4>
            <ul>
                <li><b>Click on any cell</b> to navigate to contribution details</li>
                <li>Excel comments show file contributions</li>
                <li>Visual indicators for consolidated cells</li>
            </ul>
            <h4>Verification & Quality:</h4>
            <ul>
                <li><b>Interactive comments</b> - View cell comments for detailed breakdown</li>
                <li><b>Audit reports</b> - Complete traceability of all data</li>
                <li><b>Visual indicators</b> - Orange borders on consolidated cells</li>
                <li><b>Quality checks</b> - Data validation and error detection</li>
                <li><b>Backup system</b> - Automatic file protection</li>
            </ul>
            <h4>Instructions:</h4>
            <ol>
                <li>Select a template Excel file with your desired formatting</li>
                <li>Choose a folder containing Excel files to consolidate</li>
                <li>Select where to save the consolidated file</li>
                <li>Click "Run Consolidation" to process</li>
                <li>Review the audit report for complete verification</li>
            </ol>
        """)
        info_text.setMaximumHeight(300)
        info_layout.addWidget(info_text)
        info_group.setLayout(info_layout)
        right_layout.addWidget(info_group)
        
        stats_group = QGroupBox("Statistics")
        stats_group.setStyleSheet("QGroupBox::title{background:transparent;color:#0f766e;}")
        stats_layout = QVBoxLayout()
        stats_layout.setContentsMargins(12, 12, 12, 12)
        stats_layout.setSpacing(10)
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        self.stats_text.setHtml("""
            <p>No operations performed yet.</p>
        """)
        stats_layout.addWidget(self.stats_text)
        stats_group.setLayout(stats_layout)
        right_layout.addWidget(stats_group)
        
        right_layout.addStretch()
        content_splitter.addWidget(right_widget)
        
        content_splitter.setSizes([420, 520])
        content_splitter.setStyleSheet("""
            QSplitter { border: none; }
            QSplitter::handle { background: #edf2f7; width: 6px; }
            QSplitter::handle:hover { background: #e2e8f0; }
        """)
        main_layout.addWidget(content_splitter)

        self.setLayout(main_layout)
        
        from PyQt5.QtWidgets import QShortcut
        from PyQt5.QtGui import QKeySequence
        # Auto-update shortcut removed
        
        # Auto-update shortcuts removed

    def _style_button(self, button, color, is_primary=False):
        if is_primary:
            button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    font-size: 16px;
                    font-weight: bold;
                    padding: 15px;
                    border-radius: 8px;
                }}
                QPushButton:hover {{
                    background-color: {self._adjust_color(color, -20)};
                }}
                QPushButton:pressed {{
                    background-color: {self._adjust_color(color, -40)};
                }}
                QPushButton:disabled {{
                    background-color: #cccccc;
                    color: #888888;
                }}
            """)
        else:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #f0f0f0;
                    color: #333;
                    font-size: 14px;
                    padding: 12px;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                }
                QPushButton:pressed {
                    background-color: #d0d0d0;
                }
            """)

    def _adjust_color(self, color, amount):
        """Adjust hex color by amount (-255 to 255)"""
        if color.startswith('#'):
            color = color[1:]
        rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        rgb = tuple(max(0, min(255, c + amount)) for c in rgb)
        return '#{:02x}{:02x}{:02x}'.format(*rgb)
    # ---------------- Actions ----------------
    def open_template_file(self):
        try:
            file, _ = QFileDialog.getOpenFileName(self, "Select Excel Template", "", "Excel Files (*.xlsx *.xls)")
            if file:
                self.template_path = file
                self.template_label.setText(f"📄 {os.path.basename(file)}")
                self.template_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
                self.update_stats()
        except Exception as e:
            self.report_error_with_context(
                type(e), e, e.__traceback__, 
                triggered_by="Open Template File Button",
                user_file=file if 'file' in locals() else None
            )

    def open_excel_folder(self):
        try:
            folder = QFileDialog.getExistingDirectory(self, "Select Folder Containing Excel Files")
            if folder:
                self.excel_folder = folder
                files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]
                self.folder_label.setText(f"📂 {os.path.basename(folder)}\n({len(files)} Excel files found)")
                self.folder_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
                self.update_stats()
        except Exception as e:
            self.report_error_with_context(
                type(e), e, e.__traceback__, 
                triggered_by="Open Excel Folder Button",
                user_file=folder if 'folder' in locals() else None
            )

    def open_save_folder(self):
        try:
            folder = QFileDialog.getExistingDirectory(self, "Select Save Folder")
            if folder:
                self.save_folder = folder
                self.save_label.setText(f"💾 {os.path.basename(folder)}")
                self.save_label.setStyleSheet("color: #0f766e; background: transparent; padding: 0; border-radius: 0;")
                self.update_stats()
        except Exception as e:
            self.report_error_with_context(
                type(e), e, e.__traceback__, 
                triggered_by="Open Save Folder Button",
                user_file=folder if 'folder' in locals() else None
            )

    def update_stats(self):
        stats_html = "<h3>Current Selection</h3>"
        
        if self.template_path:
            stats_html += f"<p>📄 Template: <b>{os.path.basename(self.template_path)}</b></p>"
        else:
            stats_html += "<p>📄 Template: <span style='color: #D32F2F;'>Not selected</span></p>"
            
        if self.excel_folder:
            files = [f for f in os.listdir(self.excel_folder) if f.endswith(".xlsx") and not f.startswith("~$")]
            stats_html += f"<p>📂 Source Folder: <b>{os.path.basename(self.excel_folder)}</b> ({len(files)} files)</p>"
        else:
            stats_html += "<p>📂 Source Folder: <span style='color: #D32F2F;'>Not selected</span></p>"
            
        if self.save_folder:
            stats_html += f"<p>💾 Save Location: <b>{os.path.basename(self.save_folder)}</b></p>"
        else:
            stats_html += "<p>💾 Save Location: <span style='color: #D32F2F;'>Not selected</span></p>"
            
        self.stats_text.setHtml(stats_html)

    def open_advanced_settings(self):
        """Open advanced settings dialog"""
        dialog = AdvancedConfigDialog(self)
        
        # If we have previous settings, apply them
        if self.advanced_settings:
            # Apply previous settings to dialog (would need methods to set values)
            pass
        
        if dialog.exec_() == QDialog.Accepted:
            self.advanced_settings = dialog.get_settings()
            
            # Update UI to show advanced settings are configured
            self.advanced_btn.setText("Advanced Settings")
            self.advanced_btn.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
            self.advanced_btn.setObjectName("SuccessButton")
            
            # Show summary of settings in stats
            self.update_stats_with_settings()

    def update_stats_with_settings(self):
        """Update stats display to show current advanced settings"""
        if self.advanced_settings:
            stats_html = "<h3>Current Selection</h3>"
            
            if self.template_path:
                stats_html += f"<p>📄 Template: <b>{os.path.basename(self.template_path)}</b></p>"
            else:
                stats_html += "<p>📄 Template: <span style='color: #D32F2F;'>Not selected</span></p>"
                
            if self.excel_folder:
                files = [f for f in os.listdir(self.excel_folder) if f.endswith(".xlsx") and not f.startswith("~$")]
                stats_html += f"<p>📂 Source Folder: <b>{os.path.basename(self.excel_folder)}</b> ({len(files)} files)</p>"
            else:
                stats_html += "<p>📂 Source Folder: <span style='color: #D32F2F;'>Not selected</span></p>"
                
            if self.save_folder:
                stats_html += f"<p>💾 Save Location: <b>{os.path.basename(self.save_folder)}</b></p>"
            else:
                stats_html += "<p>💾 Save Location: <span style='color: #D32F2F;'>Not selected</span></p>"
            
            # Add advanced settings summary
            stats_html += "<hr><h4>Advanced Settings Active</h4>"
            
            # File format support
            formats = []
            if self.advanced_settings['file_handling']['support_xls']:
                formats.append("XLS")
            if self.advanced_settings['file_handling']['support_csv']:
                formats.append("CSV")
            formats.append("XLSX")  # Always supported
            stats_html += f"<p>📊 Formats: {', '.join(formats)}</p>"
            
            # Data processing
            if self.advanced_settings['data_processing']['auto_convert_text']:
                stats_html += "<p>🔄 Auto-convert text numbers: Enabled</p>"
            
            # Validation
            validations = []
            if self.advanced_settings['validation']['validate_structure']:
                validations.append("Structure")
            if self.advanced_settings['validation']['validate_data_types']:
                validations.append("Data Types")
            if self.advanced_settings['validation']['validate_ranges']:
                validations.append("Value Ranges")
            
            if validations:
                stats_html += f"<p>✅ Validation: {', '.join(validations)}</p>"
            
            # Backup
            if self.advanced_settings['performance']['create_backup']:
                stats_html += "<p>💾 Backup: Enabled</p>"
                
            self.stats_text.setHtml(stats_html)
        else:
            self.update_stats()

    def run_processing(self):
        try:
            if not self.template_path or not self.excel_folder or not self.save_folder:
                missing_items = []
                if not self.template_path:
                    missing_items.append("• Select a template file")
                if not self.excel_folder:
                    missing_items.append("• Choose source folder with Excel files")
                if not self.save_folder:
                    missing_items.append("• Choose output folder for consolidated file")
                
                QMessageBox.warning(self, "Missing Information", 
                                   f"Please complete all required steps before running consolidation:\n\n"
                                   f"{chr(10).join(missing_items)}\n\n"
                                   f"All three steps are required for the consolidation to work properly.")
                return

            # Disable run button during processing
            self.run_btn.setEnabled(False)

            # Modern loading dialog
            self.loading_dialog = ModernLoadingDialog(self)
            
            # Start background worker with advanced settings
            worker_cls = globals().get('ConsolidationWorker')
            if worker_cls is None:
                self.loading_dialog.close()
                self.run_btn.setEnabled(True)
                QMessageBox.information(self, "Processing Disabled",
                                        "The consolidation worker was removed from this build,"
                                        " so processing cannot run.")
                return
            
            self.worker = worker_cls(
                self.template_path, 
                self.excel_folder, 
                self.save_folder, 
                self.advanced_settings,
                self.error_reporter,
                self.exclude_zero_percent.isChecked()
            )
            self.worker.progress.connect(self.loading_dialog.update_progress)
            self.worker.file_processed.connect(self.loading_dialog.add_processed_file)
            self.worker.finished.connect(self.on_finished)
            self.worker.start()
            
            self.loading_dialog.exec_()
            
        except Exception as e:
            # Report error with context
            self.report_error_with_context(
                type(e), e, e.__traceback__, 
                triggered_by="Run Processing Button",
                user_file=self.template_path if self.template_path else None
            )
            # Re-enable button
            if hasattr(self, 'run_btn'):
                self.run_btn.setEnabled(True)
            # Close dialog if open
            if hasattr(self, 'loading_dialog'):
                self.loading_dialog.close()

    def on_finished(self, status, message):
        self.loading_dialog.close()
        self.run_btn.setEnabled(True)
        
        if status == "success":
            # Parse the message to get file paths
            paths = message.split("|")
            consolidated_path = paths[0]
            audit_path = paths[1] if len(paths) > 1 else None
            
            # Success message with a professional dialog
            success_dialog = QDialog(self)
            success_dialog.setWindowTitle("Consolidation complete")
            success_dialog.setModal(True)
            success_dialog.setFixedSize(560, 340)
            success_dialog.setWindowFlags(success_dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)

            # Dialog-wide styling
            success_dialog.setStyleSheet(
                """
                QDialog { background: #ffffff; }
                QLabel#HeaderTitle { font-size: 16px; font-weight: 600; color: #1b5e20; }
                QLabel#BodyText { font-size: 12px; color: #2c2c2c; }
                QFrame#HeaderBar { background-color: #E8F5E9; border-radius: 6px; }
                QFrame#Separator { background-color: #e0e0e0; max-height: 1px; min-height: 1px; }
                QPushButton[primary="true"] {
                    background-color: #2e7d32; color: #ffffff; font-weight: 600; padding: 8px 14px; border-radius: 6px; border: none;
                }
                QPushButton[primary="true"]:hover { background-color: #276a2a; }
                QPushButton[secondary="true"] {
                    background-color: #eeeeee; color: #333333; font-weight: 600; padding: 8px 14px; border-radius: 6px; border: 1px solid #d5d5d5;
                }
                QPushButton[secondary="true"]:hover { background-color: #e6e6e6; }
                """
            )

            layout = QVBoxLayout(success_dialog)
            layout.setContentsMargins(18, 18, 18, 18)
            layout.setSpacing(14)

            # Header banner
            header = QFrame()
            header.setObjectName("HeaderBar")
            header_layout = QHBoxLayout(header)
            header_layout.setContentsMargins(14, 12, 14, 12)
            header_layout.setSpacing(12)

            header_icon = QLabel("✔")
            header_icon.setAlignment(Qt.AlignCenter)
            header_icon.setFixedSize(28, 28)
            header_icon.setStyleSheet(
                "background-color: #2e7d32; border-radius: 14px; color: #ffffff; font-size: 16px; font-weight: 700;"
            )
            header_layout.addWidget(header_icon)

            header_title = QLabel("Consolidation completed successfully")
            header_title.setObjectName("HeaderTitle")
            header_layout.addWidget(header_title, 1)

            layout.addWidget(header)

            # Details
            body_text = ""
            body_text += f"Consolidated file: {os.path.basename(consolidated_path)}\n"
            if audit_path:
                body_text += f"Audit report: {os.path.basename(audit_path)}\n"
            body_text += "\nTip: In Excel, click on any consolidated cell to navigate to its contribution details. Consolidated cells have orange borders and interactive comments showing the detailed breakdown of file contributions."

            message_label = QLabel(body_text)
            message_label.setObjectName("BodyText")
            message_label.setWordWrap(True)
            layout.addWidget(message_label)

            # Separator
            separator = QFrame()
            separator.setObjectName("Separator")
            layout.addWidget(separator)

            # Action buttons (right-aligned)
            btn_layout = QHBoxLayout()
            btn_layout.addStretch(1)

            consolidated_btn = QPushButton("Open consolidated file")
            consolidated_btn.setProperty("primary", True)
            consolidated_btn.clicked.connect(lambda: webbrowser.open(consolidated_path))

            if audit_path:
                audit_btn = QPushButton("Open audit report")
                audit_btn.setProperty("secondary", True)
                audit_btn.clicked.connect(lambda: webbrowser.open(audit_path))

            open_folder_btn = QPushButton("Open containing folder")
            open_folder_btn.setProperty("secondary", True)
            open_folder_btn.clicked.connect(lambda: webbrowser.open(os.path.dirname(consolidated_path)))

            close_btn = QPushButton("Close")
            close_btn.setProperty("secondary", True)
            close_btn.clicked.connect(success_dialog.accept)

            btn_layout.addWidget(open_folder_btn)
            if audit_path:
                btn_layout.addWidget(audit_btn)
            btn_layout.addWidget(consolidated_btn)
            btn_layout.addWidget(close_btn)
            layout.addLayout(btn_layout)

            # Default action
            consolidated_btn.setDefault(True)

            success_dialog.exec_()
        else:
            QMessageBox.critical(self, "Error", f"An error occurred during processing:\n\n{message}")
    


def build_global_stylesheet():
    """Return an application-wide stylesheet to provide a consistent, modern UI.
    Uses neutral surfaces, accessible contrasts, and consistent paddings.
    """
    return """
    QWidget {
        font-family: 'Segoe UI', Arial, sans-serif;
        font-size: 10pt;
        color: #1f1f1f;
    }
    QMainWindow, QWidget, QDialog {
        background-color: #deeaee; /* match app bg */
    }
    QLabel[hint="true"] { color: #616161; }

    /* App header */
    QFrame#AppHeader {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffffff, stop:1 #f7f9fb);
        border-bottom: 1px solid #e6e6e6;
    }

    /* Group containers */
    QGroupBox {
        border: 1px solid #e5e7eb;
        border-radius: 10px;
        margin-top: 14px; /* space for title */
        padding: 12px 10px 12px 10px; /* lower title a bit */
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                   stop:0 #ffffff, stop:1 #fbfbfb);
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        left: 16px;
        padding: 2px 6px; /* slight top padding to lower text */
        color: #005423;
        background: transparent; /* keep floating label transparent */
        font-weight: 600;
    }

    /* Buttons */
    QPushButton {
        background-color: #f0f0f0;
        color: #222222;
        padding: 8px 12px;
        border: 1px solid #d5d5d5;
        border-radius: 6px;
    }
    QPushButton:hover { background-color: #eaeaea; }
    QPushButton:pressed { background-color: #e0e0e0; }
    QPushButton:disabled { color: #9e9e9e; background-color: #f7f7f7; border-color: #e8e8e8; }

    /* Primary buttons can opt-in via property */
    QPushButton[primary="true"] {
        background-color: #2e7d32; color: #ffffff; border: none; font-weight: 600;
    }
    QPushButton[primary="true"]:hover { background-color: #276a2a; }
    QPushButton[primary="true"]:pressed { background-color: #225b24; }

    /* Inputs */
    QLineEdit, QTextEdit, QPlainTextEdit, QComboBox, QListWidget, QTreeWidget, QTableView {
        background-color: #fafafa;
        border: 1px solid #dcdcdc;
        border-radius: 6px;
        selection-background-color: #cce8cc;
        selection-color: #1b1b1b;
    }
    QLineEdit:focus, QTextEdit:focus, QPlainTextEdit:focus, QComboBox:focus {
        border: 1px solid #4CAF50;
        background-color: #ffffff;
    }

    /* Headers / subtle separators */
    QFrame[role="separator"], QFrame#Separator { background-color: #e0e0e0; max-height: 1px; min-height: 1px; }

    /* Scrollbars (compact, unobtrusive) */
    QScrollBar:vertical { width: 10px; background: transparent; }
    QScrollBar::handle:vertical { background: #cfcfcf; border-radius: 5px; min-height: 24px; }
    QScrollBar::handle:vertical:hover { background: #bdbdbd; }
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

    QToolTip { background-color: #333333; color: #ffffff; border: none; padding: 6px 8px; border-radius: 4px; }
    """


def main():
    """Main entry point for the Excel Consolidator application."""
    # Improve rendering on HiDPI displays and reduce pixelation
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    # Apply global stylesheet
    try:
        app.setStyleSheet(build_global_stylesheet())
    except Exception:
        pass
    
    # Update app icon loading
    window = ExcelProcessorApp()
    app.setWindowIcon(window.windowIcon())
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()