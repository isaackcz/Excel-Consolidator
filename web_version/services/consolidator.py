"""
Excel Consolidation Service - Extracted from desktop app
Full performance and accuracy functions - NO advanced settings UI
All core processing logic from desktop ConsolidationWorker
"""
import os
import glob
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, Font, PatternFill
from decimal import Decimal, InvalidOperation
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelConsolidator:
    """
    Consolidates multiple Excel files into one template file
    Full accuracy and performance from desktop application
    """
    
    def __init__(self, template_path, source_folder, settings=None, progress_callback=None):
        """
        Initialize consolidator
        
        Args:
            template_path: Path to Excel template file
            source_folder: Folder containing source Excel files
            settings: Dict of processing settings
            progress_callback: Function(current, total, filename) for progress updates
        """
        self.template_path = template_path
        self.source_folder = source_folder
        self.settings = settings or {}
        self.progress_callback = progress_callback
        
        # Processing flags
        self.convert_text_to_numbers = self.settings.get('convert_text_to_numbers', True)
        self.convert_percentages = self.settings.get('convert_percentages', True)
        self.skip_validation = self.settings.get('skip_validation', True)
        self.exclude_zero_percent = self.settings.get('exclude_zero_percent', False)
        
        logger.info(f"Consolidator initialized: template={template_path}, sources={source_folder}")
    
    # ============================================================================
    # FORMAT DETECTION METHODS (from desktop app)
    # ============================================================================
    
    def _is_percentage_format(self, format_str: str) -> bool:
        """Enhanced percentage format detection with comprehensive patterns."""
        if not format_str:
            return False
        
        format_str = str(format_str).lower()
        percentage_patterns = [
            '%', 'percent', '0.0%', '0.00%', '0%', '#,##0%', '#,##0.0%', '#,##0.00%',
            '0.0%', '0.00%', '0%', '0.0%', '0.00%', '0%', '0.0%', '0.00%',
            'general%', 'standard%', 'percentage', 'pct', 'pct%'
        ]
        
        return any(pattern in format_str for pattern in percentage_patterns)
    
    def _is_currency_format(self, format_str: str) -> bool:
        """Enhanced currency format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str)
        currency_symbols = ['$', '€', '£', '¥', '₽', '₹', '₩', '₪', '₦', '₡', '₨', '₫', '₱', '₲', '₴', '₵', '₸', '₼', '₾', '₿']
        currency_patterns = ['currency', 'money', 'dollar', 'euro', 'pound', 'yen']
        
        return (any(symbol in format_str for symbol in currency_symbols) or 
                any(pattern in format_str.lower() for pattern in currency_patterns))
    
    def _is_number_format(self, format_str: str) -> bool:
        """Enhanced number format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str)
        
        # First check if it's already identified as percentage or currency
        if self._is_percentage_format(format_str) or self._is_currency_format(format_str):
            return False
        
        number_patterns = [
            '0.00', '#,##0', '0.0', '0', '#,##0.00', '#,##0.0', '#,##0',
            'general', 'standard', 'number', 'numeric', 'decimal',
            '0.000', '0.0000', '#,##0.000', '#,##0.0000'
        ]
        
        return any(pattern in format_str for pattern in number_patterns)
    
    def _is_date_format(self, format_str: str) -> bool:
        """Enhanced date format detection."""
        if not format_str:
            return False
        
        format_str = str(format_str).lower()
        date_patterns = [
            'mm/dd/yyyy', 'dd/mm/yyyy', 'yyyy-mm-dd', 'mm-dd-yyyy', 'dd-mm-yyyy',
            'mm/dd/yy', 'dd/mm/yy', 'yy-mm-dd', 'mm-dd-yy', 'dd-mm-yy',
            'm/d/yyyy', 'd/m/yyyy', 'm/d/yy', 'd/m/yy',
            'date', 'time', 'datetime', 'timestamp'
        ]
        
        return any(pattern in format_str for pattern in date_patterns)
    
    def _get_consolidation_method(self, format_info: dict) -> str:
        """Determine the appropriate consolidation method based on cell format."""
        if format_info.get('is_percentage', False):
            return 'average'
        elif format_info.get('is_currency', False):
            return 'sum'
        elif format_info.get('is_number', False):
            return 'sum'
        elif format_info.get('is_date', False):
            return 'sum'
        else:
            return 'sum'
    
    def _is_total_cell(self, cell) -> bool:
        """Detect if a cell is likely a total row/column based on common patterns."""
        if cell.value is None:
            return False
        
        # Check for common total indicators in cell value
        value_str = str(cell.value).lower().strip()
        total_indicators = ['total', 'sum', 'subtotal', 'grand total', 'totaal', 'gesamt']
        
        # Check if cell value contains total indicators
        if any(indicator in value_str for indicator in total_indicators):
            return True
        
        return False
    
    # ============================================================================
    # VALUE PROCESSING METHODS (from desktop app)
    # ============================================================================
    
    def _process_cell_value_with_format_verification(self, value, format_info, coord, file_label, stop_on_error=False):
        """
        Process cell value with comprehensive format verification.
        Ensures data is converted according to template format requirements.
        
        Args:
            value: The cell value to process
            format_info: Dictionary containing format information from template
            coord: Cell coordinate (e.g., 'A1')
            file_label: Name of the source file
            stop_on_error: Whether to stop on format mismatches
            
        Returns:
            Decimal value or None if processing failed
        """
        if value is None or value == "":
            return None
            
        # Handle different format types based on template requirements
        if format_info.get('is_percentage', False):
            return self._process_percentage_value(value, coord, file_label, stop_on_error)
        elif format_info.get('is_currency', False):
            return self._process_currency_value(value, coord, file_label, stop_on_error)
        elif format_info.get('is_number', False):
            return self._process_number_value(value, coord, file_label, stop_on_error)
        else:
            # Default processing for unformatted cells
            return self._process_default_value(value, coord, file_label, stop_on_error)
    
    def _process_percentage_value(self, value, coord, file_label, stop_on_error):
        """Process percentage values with strict format verification."""
        try:
            # Handle different percentage input formats
            if isinstance(value, (int, float)):
                # Normalize numeric inputs to PERCENT POINTS for averaging
                # Rules:
                #  - Values > 1 are treated as percent points (e.g., 82.5 means 82.5%)
                #  - Values between 0 and 1 are decimals; convert to percent points (0.825 → 82.5)
                numeric_val = float(value)
                if 0 <= numeric_val <= 1:
                    normalized = numeric_val * 100.0
                else:
                    normalized = numeric_val
                return Decimal(str(normalized))
            elif isinstance(value, str):
                text = str(value).strip().replace(",", "")
                if text.endswith('%'):
                    # Remove % and interpret as percent points directly
                    val = Decimal(text[:-1])
                    return val
                else:
                    # Parse as number; apply same normalization as numeric path
                    numeric_val = float(text)
                    if 0 <= numeric_val <= 1:
                        normalized = numeric_val * 100.0
                    else:
                        normalized = numeric_val
                    return Decimal(str(normalized))
            else:
                return None
        except Exception as e:
            logger.warning(f"Could not process percentage value at {coord}: {value}")
            return None
    
    def _process_currency_value(self, value, coord, file_label, stop_on_error):
        """Process currency values with format verification."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Remove currency symbols and parse
                text = str(value).strip()
                for symbol in ['$', '€', '£', '¥', '₽', '₹', '₩', '₪', '₦', '₡', '₨', '₫', '₱']:
                    text = text.replace(symbol, '')
                text = text.replace(',', '').replace(' ', '')
                return Decimal(text)
            else:
                return None
        except Exception as e:
            logger.warning(f"Could not process currency value at {coord}: {value}")
            return None
    
    def _process_number_value(self, value, coord, file_label, stop_on_error):
        """Process number values with format verification."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Remove common formatting characters
                text = str(value).strip().replace(",", "").replace(" ", "")
                return Decimal(text)
            else:
                return None
        except Exception as e:
            logger.warning(f"Could not process number value at {coord}: {value}")
            return None
    
    def _process_default_value(self, value, coord, file_label, stop_on_error):
        """Process values with default (unformatted) handling."""
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            elif isinstance(value, str):
                # Try to parse as number
                text = str(value).strip().replace(",", "")
                
                # Handle percentages if enabled
                if self.convert_percentages and text.endswith('%'):
                    try:
                        numeric = float(text[:-1])
                        return Decimal(str(numeric))
                    except:
                        pass
                
                # Try to convert text to number
                if self.convert_text_to_numbers:
                    try:
                        return Decimal(text)
                    except:
                        pass
            return None
        except Exception:
            return None
    
    def _convert_to_percentage_format(self, value, coord):
        """
        Convert any value to percentage format (decimal for Excel).
        Handles: numbers (82.5), decimals (0.825), text ("82.5%", "50"), etc.
        
        Returns: Decimal value for Excel (e.g., 0.825 for 82.5%)
        """
        try:
            # If already a number
            if isinstance(value, (int, float)):
                # Values > 1 are percentage points (82.5 means 82.5%)
                if value > 1:
                    return value / 100  # 82.5 → 0.825
                # Values 0-1 are already decimals (0.825 means 82.5%)
                elif 0 <= value <= 1:
                    return value  # 0.825 → 0.825
                else:
                    # Negative or unusual values, treat as percentage points
                    return value / 100
            
            # If text, parse it
            elif isinstance(value, str):
                text = value.strip()
                
                # Remove % symbol if present
                if text.endswith('%'):
                    # "82.5%" → 82.5 → 0.825
                    numeric = float(text[:-1].replace(',', ''))
                    return numeric / 100
                else:
                    # "82.5" or "0.825" - determine which
                    numeric = float(text.replace(',', ''))
                    if numeric > 1:
                        return numeric / 100  # 82.5 → 0.825
                    else:
                        return numeric  # 0.825 → 0.825
            
            return None
            
        except Exception as e:
            logger.warning(f"⚠️ Could not convert {coord} value '{value}' to percentage: {e}")
            return None
    
    def _convert_to_number_format(self, value, coord, is_currency=False):
        """
        Convert any value to number format.
        Handles: numbers (100), text ("100", "$100", "1,234"), etc.
        
        Returns: Numeric value
        """
        try:
            # If already a number, return as-is
            if isinstance(value, (int, float)):
                return value
            
            # If text, parse it
            elif isinstance(value, str):
                text = value.strip()
                
                # Remove currency symbols
                for symbol in ['$', '€', '£', '¥', '₽', '₹', '₩', '₪', '₦', '₡', '₨', '₫', '₱']:
                    text = text.replace(symbol, '')
                
                # Remove commas and spaces
                text = text.replace(',', '').replace(' ', '')
                
                # Remove % symbol if present (shouldn't be here, but handle it)
                if text.endswith('%'):
                    text = text[:-1]
                
                # Parse to number
                return float(text)
            
            return None
            
        except Exception as e:
            logger.warning(f"⚠️ Could not convert {coord} value '{value}' to number: {e}")
            return None
    
    # ============================================================================
    # MAIN CONSOLIDATION LOGIC
    # ============================================================================
    
    def consolidate(self):
        """
        Main consolidation method with full desktop app logic
        Returns: Path to consolidated output file
        """
        logger.info("=" * 60)
        logger.info("Starting Excel Consolidation (Full Desktop Logic)")
        logger.info("=" * 60)
        
        # Find all Excel files
        files = self._get_excel_files()
        if not files:
            raise ValueError("No Excel files found in source folder")
        
        logger.info(f"Found {len(files)} Excel files to consolidate")
        
        # Load template with full analysis
        logger.info("📋 Loading and analyzing template...")
        template_wb = openpyxl.load_workbook(self.template_path, data_only=False, read_only=False)
        output_ws = template_wb.active
        
        logger.info(f"Template worksheet loaded: {output_ws.title}")
        
        # Storage for totals and contributions
        totals = {}
        contributions = {}
        percent_counts = {}
        
        # Analyze template for format information (CRITICAL for accuracy)
        logger.info("🔍 Analyzing template cell formats...")
        coord_format_info, template_coords = self._analyze_template_formats_enhanced(output_ws)
        
        # Log percentage cells found for debugging
        percent_cells = [coord for coord, info in coord_format_info.items() if info.get('is_percentage')]
        if percent_cells:
            logger.info(f"📊 Percentage cells detected: {percent_cells[:5]}{'...' if len(percent_cells) > 5 else ''}")
        
        total_files_count = len(files)
        
        # Process each source file
        logger.info(f"📁 Processing {total_files_count} files...")
        for idx, file in enumerate(files, 1):
            if self.progress_callback:
                self.progress_callback(idx, len(files), os.path.basename(file))
            
            logger.info(f"Processing [{idx}/{len(files)}]: {os.path.basename(file)}")
            
            try:
                self._process_file_enhanced(
                    file,
                    totals,
                    contributions,
                    percent_counts,
                    coord_format_info,
                    template_coords,
                    total_files_count,
                    idx
                )
            except Exception as e:
                logger.error(f"Error processing {file}: {str(e)}")
                # Continue with next file
        
        # Write consolidated values to template (with full desktop logic)
        logger.info("✍️ Writing consolidated values to template...")
        self._write_consolidated_values_enhanced(
            output_ws,
            totals,
            contributions,
            percent_counts,
            coord_format_info,
            total_files_count
        )
        
        # Save output file
        output_path = self._generate_output_path()
        logger.info(f"💾 Saving consolidated file: {output_path}")
        template_wb.save(output_path)
        template_wb.close()
        
        logger.info(f"✅ Consolidation complete: {output_path}")
        logger.info("=" * 60)
        
        return output_path
    
    def _get_excel_files(self):
        """Get list of Excel files from source folder"""
        files = []
        
        # Look for .xlsx files
        xlsx_files = glob.glob(os.path.join(self.source_folder, '*.xlsx'))
        files.extend([f for f in xlsx_files if not os.path.basename(f).startswith('~$')])
        
        # Optionally include .xls files
        if self.settings.get('support_xls', True):
            xls_files = glob.glob(os.path.join(self.source_folder, '*.xls'))
            files.extend([f for f in xls_files if not os.path.basename(f).startswith('~$')])
        
        return sorted(files)
    
    def _analyze_template_formats_enhanced(self, worksheet):
        """
        Enhanced template format analysis with comprehensive cell format verification
        Exactly as in desktop app for maximum accuracy
        CRITICAL: Also creates template_coords set for filtering source file cells
        """
        format_info = {}
        template_coords = set()  # CRITICAL: Track all template cell coordinates
        cell_count = 0
        processed_cells = 0
        
        logger.info("🔍 Analyzing template cells for format detection...")
        
        for row in worksheet.iter_rows():
            for cell in row:
                coord = cell.coordinate
                template_coords.add(coord)  # CRITICAL: Add EVERY coord to set (matches desktop app line 1892)
                cell_count += 1
                
                # Process cells for format info (but add ALL coords to set above)
                # Skip completely empty cells for performance in format detection
                if cell.value is None and not cell.number_format:
                    continue
                
                processed_cells += 1
                number_format = cell.number_format or ''
                
                # Comprehensive format detection
                is_percentage = self._is_percentage_format(number_format)
                is_currency = self._is_currency_format(number_format)
                is_number = self._is_number_format(number_format)
                is_date = self._is_date_format(number_format)
                
                info = {
                    'number_format': number_format,
                    'is_percentage': is_percentage,
                    'is_currency': is_currency,
                    'is_number': is_number,
                    'is_date': is_date,
                    'consolidation_method': 'average' if is_percentage else 'sum'
                }
                
                format_info[coord] = info
        
        logger.info(f"Analyzed {processed_cells} cells out of {cell_count} total cells in template")
        logger.info(f"Template coordinates tracked: {len(template_coords)}")
        logger.info(f"Percentage cells: {len([c for c, i in format_info.items() if i.get('is_percentage')])}")
        logger.info(f"Currency cells: {len([c for c, i in format_info.items() if i.get('is_currency')])}")
        logger.info(f"Number cells: {len([c for c, i in format_info.items() if i.get('is_number')])}")
        
        # Return BOTH format_info AND template_coords (matches desktop app)
        return format_info, template_coords
    
    def _process_file_enhanced(self, filepath, totals, contributions, percent_counts, 
                              coord_format_info, template_coords, total_files, file_idx):
        """
        Enhanced file processing with full desktop app logic
        """
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        
        file_label = os.path.splitext(os.path.basename(filepath))[0]
        
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                coord = cell.coordinate
                
                # If template coords specified, skip cells not in template
                if template_coords and coord not in template_coords:
                    continue
                
                # Skip empty cells
                if value is None or value == '':
                    continue
                
                # CRITICAL: Skip formulas in SOURCE files to prevent double-counting
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    logger.debug(f"⏩ Skipping formula cell {coord} in {file_label}")
                    continue
                
                # Get format info
                format_info = coord_format_info.get(coord, {})
                
                # Process cell value with format verification
                val = self._process_cell_value_with_format_verification(
                    value, format_info, coord, file_label, False
                )
                if val is None:
                    continue
                
                # Determine consolidation method
                consolidation_method = format_info.get('consolidation_method', 'sum')
                
                if consolidation_method == 'average':
                    # Percentage cells: accumulate for average calculation
                    current_total = totals.get(coord)
                    totals[coord] = (current_total + val) if current_total is not None else val
                    
                    # Initialize count to total files on first encounter
                    if coord not in percent_counts:
                        if self.exclude_zero_percent:
                            # When excluding zeros: only count files with non-zero values
                            percent_counts[coord] = 0
                        else:
                            # Default: count all files (including files with 0% values)
                            percent_counts[coord] = total_files
                    
                    # If excluding zeros, increment count only for non-zero values
                    if self.exclude_zero_percent and val != 0:
                        percent_counts[coord] += 1
                    
                    logger.debug(f"📊 Percentage cell {coord}: {val} (from {file_label})")
                    
                else:
                    # Sum values
                    current_total = totals.get(coord)
                    totals[coord] = (current_total + val) if current_total is not None else val
                    
                    logger.debug(f"🔢 Sum cell {coord}: {val} (from {file_label})")
                
                # Track contributions
                if coord not in contributions:
                    contributions[coord] = {}
                prev = contributions[coord].get(file_label)
                contributions[coord][file_label] = (prev + val) if prev is not None else val
        
        wb.close()
    
    def _write_consolidated_values_enhanced(self, worksheet, totals, contributions, 
                                           percent_counts, coord_format_info, total_files):
        """
        Enhanced value writing with full desktop app logic and formatting
        """
        # Orange border for consolidated cells
        thin_orange = Border(
            left=Side(style='thin', color='FF8C00'),
            right=Side(style='thin', color='FF8C00'),
            top=Side(style='thin', color='FF8C00'),
            bottom=Side(style='thin', color='FF8C00')
        )
        
        for coord, value in totals.items():
            cell = worksheet[coord]
            
            if isinstance(cell, MergedCell):
                continue
            
            format_info = coord_format_info.get(coord, {})
            consolidation_method = format_info.get('consolidation_method', 'sum')
            
            try:
                if consolidation_method == 'average':
                    # Calculate average for percentage cells
                    count = max(1, percent_counts.get(coord, 1))
                    avg_value = float(value / Decimal(count))
                    
                    # Excel expects percentages as decimals (e.g., 0.825 for 82.5%)
                    cell.value = avg_value / 100
                    
                    # Maintain percentage format
                    template_format = format_info.get('number_format', '0.00%')
                    cell.number_format = template_format
                    
                    logger.info(f"✅ {coord}: Average = {avg_value:.2f}% (format: {template_format})")
                    
                else:
                    # Sum for other cells
                    cell.value = float(value)
                    
                    # Apply formatting based on cell type
                    if format_info.get('is_currency', False):
                        template_format = format_info.get('number_format', '$#,##0.00')
                        cell.number_format = template_format
                        logger.debug(f"✅ {coord}: Currency sum = {float(value):,.2f}")
                        
                    elif format_info.get('is_number', False):
                        template_format = format_info.get('number_format', '#,##0.00')
                        cell.number_format = template_format
                        logger.debug(f"✅ {coord}: Number sum = {float(value):,.2f}")
                    
            except Exception as e:
                logger.error(f"Error writing value to {coord}: {e}")
                cell.value = float(value) if value is not None else 0
            
            # Add comment showing contributions
            file_map = contributions.get(coord, {})
            if file_map:
                comment_text = self._build_comment_text_enhanced(
                    coord, value, file_map, format_info, percent_counts, total_files
                )
                cell.comment = Comment(comment_text, "Excel Consolidator Web")
            
            # Add orange border to indicate consolidated cell
            cell.border = thin_orange
    
    def _build_comment_text_enhanced(self, coord, total_value, file_map, format_info, 
                                    percent_counts, total_files):
        """Build enhanced comment text showing file contributions"""
        items = sorted(file_map.items(), key=lambda x: x[0].lower())
        max_name = max((len(n) for n, _ in items), default=4)
        
        lines = []
        lines.append("Consolidation Summary")
        lines.append(f"Cell: {coord}")
        
        # Enhanced summary based on cell format
        is_percent = format_info.get('is_percentage', False)
        
        if is_percent:
            count = max(1, int(percent_counts.get(coord, 1)))
            avg_val = (total_value / Decimal(count))
            num_contributors = len([v for v in file_map.values() if v != 0])
            
            if self.exclude_zero_percent:
                # Excluding zeros: count only includes files with non-zero values
                lines.append(f"Average: {float(avg_val):,.2f}% (from {count} files with values")
                if num_contributors != count:
                    lines.append(f", {num_contributors} non-zero")
                lines.append(", zero values excluded)\n")
            else:
                # Default: count includes ALL files
                lines.append(f"Average: {float(avg_val):,.2f}% (from {count} files")
                if num_contributors < count:
                    lines.append(f", {num_contributors} with values, {count - num_contributors} empty")
                lines.append(")\n")
        elif format_info.get('is_currency', False):
            lines.append(f"Total: ${float(total_value):,.2f}\n")
        elif format_info.get('is_number', False):
            lines.append(f"Total: {float(total_value):,.2f}\n")
        else:
            lines.append(f"Total: {float(total_value):,.2f}\n")
        
        lines.append("Contributors (file  |  value)")
        lines.append("-" * max(26, max_name + 10))
        
        # List each file's contribution
        for name, v in items:
            pad = " " * (max_name - len(name))
            if format_info.get('is_percentage', False):
                lines.append(f"{name}{pad}  |  {float(v):.2f}%")
            elif format_info.get('is_currency', False):
                lines.append(f"{name}{pad}  |  ${float(v):,.2f}")
            else:
                lines.append(f"{name}{pad}  |  {float(v):,.2f}")
        
        return "\n".join(lines)
    
    def _generate_output_path(self):
        """Generate output file path with timestamp"""
        # Use same folder as template
        template_dir = os.path.dirname(self.template_path)
        template_ext = os.path.splitext(self.template_path)[1]
        
        # Generate filename with date
        date_str = datetime.now().strftime("%b_%d_%Y")
        filename = f"Consolidated_{date_str}{template_ext}"
        
        return os.path.join(template_dir, filename)